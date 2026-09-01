"""Farm Reports: Heifers To Scan from inventory."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from app.models import Base, HerdInventory
from app.services.farm_reports import (
    BLANK_PEN,
    ETAG4_COL_WIDTH_MM,
    ETAG4_XLSX_COL_WIDTH,
    PDF_ROWS_PER_PAGE,
    apply_pen_filter,
    build_heifers_to_scan_pdf,
    build_heifers_to_scan_xlsx,
    build_report_pdf,
    build_report_xlsx,
    collars_to_put_on,
    etag4,
    etag4_column_grid,
    etag4_pdf_col_widths,
    farm_reports,
    heifers_to_scan,
    heifers_to_scan_and_collars,
)


def _db() -> Session:
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(bind=engine)
    return sessionmaker(bind=engine, autoflush=False, autocommit=False)()


def _heifer(
    *,
    farm: str = "CM",
    cow_id: str,
    etag: str | None = "UK740651324400",
    category: str = "Youngstock",
    rc: float | None = 3,
    dslh: float | None = 40,
    rpro: str | None = "BRED",
    remark: str | None = "BRED",
    pen: str | None = "12",
    aged: int | None = 420,
    tbrd: int | None = 1,
    ewgt: float | None = None,
    lact: float | None = 0,
    httag: str | None = None,
    rum: float | None = None,
) -> HerdInventory:
    return HerdInventory(
        farm=farm,
        cow_id=cow_id,
        etag=etag,
        category=category,
        rc=rc,
        dslh=dslh,
        rpro=rpro,
        remark=remark,
        pen=pen,
        aged=aged,
        tbrd=tbrd,
        ewgt=ewgt,
        lact=lact,
        httag=httag,
        rum=rum,
    )


def test_etag4_strips_spaces_and_keeps_leading_zero() -> None:
    assert etag4("UK740651324400") == "4400"
    assert etag4("UK740651324400     ") == "4400"
    assert etag4("UK 7406 5132 4400") == "4400"
    assert etag4("UK740651300111") == "0111"
    assert etag4(None) is None


def test_heifers_to_scan_filters_youngstock_rc_and_dslh() -> None:
    db = _db()
    db.add_all(
        [
            _heifer(cow_id="100", rc=3, dslh=32, etag="UK740651324400     ", httag="12", rum=0),
            _heifer(cow_id="101", rc=4, dslh=50, etag="UK740651300111"),
            _heifer(cow_id="102", rc=3, dslh=31),
            _heifer(cow_id="103", rc=3, dslh=40, category="Dairy"),
            _heifer(cow_id="104", rc=2, dslh=80),
            _heifer(cow_id="105", farm="GAD", rc=3, dslh=90),
        ]
    )
    db.commit()

    result = heifers_to_scan(db, "CM")
    ids = [row["id"] for row in result["rows"]]
    assert result["count"] == 2
    assert ids == ["101", "100"]
    assert result["rows"][0]["etag4"] == "0111"
    assert result["rows"][1]["etag4"] == "4400"
    assert result["rows"][1]["dslh"] == 32
    assert result["rows"][1]["rpro"] == "BRED"
    assert result["rows"][1]["remark"] == "BRED"
    assert result["rows"][1]["pen"] == "12"
    assert result["rows"][1]["tbrd"] == 1
    assert result["rows"][1]["broken_collar"] is False


def test_heifers_to_scan_pen_filter_and_blank_pen() -> None:
    db = _db()
    db.add_all(
        [
            _heifer(cow_id="1", pen="12"),
            _heifer(cow_id="2", pen="8"),
            _heifer(cow_id="3", pen=None),
        ]
    )
    db.commit()
    result = heifers_to_scan(db, "CM")
    assert [pen["id"] for pen in result["pens"]] == ["8", "12", BLANK_PEN]
    filtered = heifers_to_scan(db, "CM", pens=["8", BLANK_PEN])
    assert filtered["count"] == 3
    assert filtered["filtered_count"] == 2
    assert {row["id"] for row in filtered["rows"]} == {"2", "3"}


def test_apply_pen_filter_empty_selection() -> None:
    rows = [{"id": "1", "pen": "12"}]
    assert apply_pen_filter(rows, ["__no_match__"]) == []


def test_farm_reports_widget_count_matches_table() -> None:
    db = _db()
    db.add(_heifer(cow_id="200", rc=4, dslh=33))
    db.commit()
    payload = farm_reports(db, "cm")
    assert payload["farm"] == "CM"
    assert payload["farm_label"] == "Cwrt Malle"
    assert payload["widgets"][0]["id"] == "heifers-to-scan"
    assert payload["widgets"][0]["count"] == 1
    assert payload["widgets"][1]["id"] == "collars-to-put-on"
    assert payload["widgets"][1]["count"] == 0
    assert payload["widgets"][2]["id"] == "heifers-to-scan-and-collars"
    assert payload["widgets"][2]["title"] == "Heifers To Scan & Collars"
    assert payload["widgets"][2]["count"] == 1
    assert payload["heifers_to_scan"]["count"] == 1
    assert payload["collars_to_put_on"]["count"] == 0
    assert payload["heifers_to_scan_and_collars"]["count"] == 1


def test_heifers_to_scan_xlsx_export() -> None:
    db = _db()
    db.add(_heifer(cow_id="1", remark="SCAN"))
    db.commit()
    content = build_heifers_to_scan_xlsx(heifers_to_scan(db, "CM"))
    assert content[:2] == b"PK"
    wb = load_workbook(BytesIO(content))
    assert wb.sheetnames == ["Heifers To Scan", "ETAG4"]
    assert wb["ETAG4"].column_dimensions["A"].width == ETAG4_XLSX_COL_WIDTH


def test_heifers_to_scan_xlsx_etag4_only() -> None:
    db = _db()
    db.add(_heifer(cow_id="1", remark="SCAN"))
    db.commit()
    content = build_heifers_to_scan_xlsx(heifers_to_scan(db, "CM"), etag4_only=True)
    wb = load_workbook(BytesIO(content))
    assert wb.sheetnames == ["ETAG4"]
    assert wb.active["A1"].value == "ETAG4"
    assert wb.active.column_dimensions["A"].width == ETAG4_XLSX_COL_WIDTH


def test_heifers_to_scan_pdf_export() -> None:
    import pytest

    pytest.importorskip("reportlab")
    db = _db()
    db.add_all(
        [
            _heifer(cow_id=str(i), pen="12" if i < 25 else "8")
            for i in range(PDF_ROWS_PER_PAGE + 5)
        ]
    )
    db.commit()
    report = heifers_to_scan(db, "CM")
    pdf = build_heifers_to_scan_pdf(report)
    assert pdf.startswith(b"%PDF")
    assert len(pdf) > 200
    assert report["count"] == PDF_ROWS_PER_PAGE + 5


def test_etag4_grid_uses_forty_row_columns() -> None:
    rows = [{"etag4": f"{i:05d}"} for i in range(200)]
    grid = etag4_column_grid(rows)
    assert len(grid) == 40
    assert len(grid[0]) == 5
    assert grid[0] == ["00000", "00040", "00080", "00120", "00160"]
    assert grid[39][0] == "00039"
    assert grid[0][4] == "00160"

    short = etag4_column_grid([{"etag4": "1"}] * 41)
    assert len(short) == 40
    assert len(short[0]) == 2
    assert short[0] == ["1", "1"]
    assert short[1][1] == ""


def test_etag4_pdf_columns_stay_compact() -> None:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm

    usable = A4[0] - 16 * mm
    widths = etag4_pdf_col_widths(5, usable)
    assert len(widths) == 5
    assert widths[0] == ETAG4_COL_WIDTH_MM * mm
    assert sum(widths) < usable

    crowded = etag4_pdf_col_widths(20, usable)
    assert abs(sum(crowded) - usable) < 0.01


def test_heifers_to_scan_pdf_etag4_only() -> None:
    import pytest

    pytest.importorskip("reportlab")
    db = _db()
    db.add(_heifer(cow_id="1"))
    db.commit()
    pdf = build_heifers_to_scan_pdf(heifers_to_scan(db, "CM"), etag4_only=True)
    assert pdf.startswith(b"%PDF")
    assert len(pdf) > 200


def test_collars_to_put_on_filters_httag_rum_rc_and_lact() -> None:
    db = _db()
    db.add_all(
        [
            _heifer(
                cow_id="200",
                rc=0,
                ewgt=385,
                aged=390,
                remark="COLLAR",
                httag="0",
                etag="UK740651300222",
            ),
            _heifer(
                cow_id="201",
                rc=0,
                ewgt=400.4,
                aged=410,
                httag=None,
                etag="UK740651300111",
            ),
            _heifer(
                cow_id="207",
                rc=3,
                ewgt=410,
                httag="12",
                rum=0,
                rpro="BRED",
                etag="UK740651300333",
            ),
            _heifer(cow_id="202", rc=0, ewgt=384.9, httag="0"),
            _heifer(cow_id="203", rc=0, ewgt=None, httag="0"),
            _heifer(cow_id="204", rc=3, ewgt=420, httag="0"),
            _heifer(cow_id="205", rc=0, ewgt=500, httag="0", category="Dairy"),
            _heifer(cow_id="206", farm="GAD", rc=0, ewgt=500, httag="0"),
            _heifer(cow_id="208", rc=0, ewgt=400, httag="18", rum=50),
            _heifer(cow_id="209", rc=2, ewgt=400, httag="18", rum=0),
            _heifer(cow_id="210", rc=0, ewgt=400, httag="0", lact=1),
            _heifer(cow_id="211", rc=4, ewgt=400, httag="22", rum=0, lact=1),
        ]
    )
    db.commit()

    result = collars_to_put_on(db, "CM")
    by_id = {row["id"]: row for row in result["rows"]}
    assert result["count"] == 3
    assert set(by_id) == {"200", "201", "207"}
    assert by_id["200"]["httag"] == 0
    assert by_id["200"]["broken_collar"] is False
    assert by_id["200"]["ewgt"] == 385
    assert by_id["200"]["aged"] == 390
    assert by_id["200"]["remark"] == "COLLAR"
    assert by_id["200"]["pen"] == "12"
    assert by_id["201"]["etag4"] == "0111"
    assert by_id["207"]["httag"] == 12
    assert by_id["207"]["broken_collar"] is True
    assert by_id["207"]["rpro"] == "BRED"
    assert [row["id"] for row in result["rows"]] == ["201", "200", "207"]


def test_collars_to_put_on_xlsx_export() -> None:
    db = _db()
    db.add(_heifer(cow_id="1", rc=0, ewgt=385, remark="ON", httag="0"))
    db.commit()
    content = build_report_xlsx(collars_to_put_on(db, "CM"))
    wb = load_workbook(BytesIO(content))
    assert wb.sheetnames == ["Collars To Put On", "ETAG4"]
    assert [cell.value for cell in wb.active[1]] == [
        "ID", "REMARK", "ETAG4", "EWGT", "HTTAG", "AGED", "RPRO", "PEN"
    ]
    assert wb.active["D2"].value == "385"


def test_collars_to_put_on_pdf_export() -> None:
    import pytest

    pytest.importorskip("reportlab")
    db = _db()
    db.add(_heifer(cow_id="1", rc=0, ewgt=400, httag="0"))
    db.commit()
    pdf = build_report_pdf(collars_to_put_on(db, "CM"))
    assert pdf.startswith(b"%PDF")
    assert len(pdf) > 200


def test_heifers_to_scan_and_collars_unions_lists_and_keeps_ewgt_floor() -> None:
    db = _db()
    db.add_all(
        [
            _heifer(
                cow_id="100",
                rc=3,
                dslh=40,
                ewgt=200,
                httag="12",
                rum=80,
                etag="UK740651300111",
            ),
            _heifer(
                cow_id="200",
                rc=0,
                dslh=10,
                ewgt=385,
                httag="0",
                etag="UK740651300222",
            ),
            _heifer(
                cow_id="201",
                rc=0,
                dslh=10,
                ewgt=384.9,
                httag="0",
                etag="UK740651300333",
            ),
            _heifer(
                cow_id="207",
                rc=3,
                dslh=50,
                ewgt=410,
                httag="12",
                rum=0,
                etag="UK740651300444",
            ),
            _heifer(
                cow_id="208",
                rc=0,
                dslh=10,
                ewgt=400,
                httag="18",
                rum=0,
                etag="UK740651300555",
            ),
        ]
    )
    db.commit()

    result = heifers_to_scan_and_collars(db, "CM")
    by_id = {row["id"]: row for row in result["rows"]}
    assert result["id"] == "heifers-to-scan-and-collars"
    assert result["title"] == "Heifers To Scan & Collars"
    assert result["count"] == 4
    assert set(by_id) == {"100", "200", "207", "208"}
    assert "201" not in by_id
    assert by_id["100"]["dslh"] == 40
    assert by_id["100"]["ewgt"] == 200
    assert by_id["100"]["broken_collar"] is False
    assert by_id["100"]["reason"] == "Preg Check"
    assert by_id["200"]["ewgt"] == 385
    assert by_id["200"]["httag"] == 0
    assert by_id["200"]["reason"] == "Put Collar On"
    assert by_id["207"]["broken_collar"] is True
    assert by_id["207"]["dslh"] == 50
    assert by_id["207"]["reason"] == "Preg Check & Faulty Collar"
    assert by_id["208"]["broken_collar"] is True
    assert by_id["208"]["reason"] == "Faulty Collar"
    assert [row["id"] for row in result["rows"]] == ["100", "200", "207", "208"]


def test_heifers_to_scan_and_collars_xlsx_export() -> None:
    db = _db()
    db.add(_heifer(cow_id="1", rc=3, dslh=40, ewgt=200))
    db.commit()
    content = build_report_xlsx(heifers_to_scan_and_collars(db, "CM"))
    wb = load_workbook(BytesIO(content))
    assert wb.sheetnames == ["Heifers To Scan & Collars", "ETAG4"]
    assert [cell.value for cell in wb.active[1]] == [
        "ID", "REMARK", "REASON", "ETAG4", "DSLH", "TBRD", "EWGT", "HTTAG", "AGED", "RPRO", "PEN"
    ]
