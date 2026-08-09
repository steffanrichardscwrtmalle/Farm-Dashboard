"""Parse the milk haulier's emailed XLSX collection report.

The sheet lists loads grouped under date headers in column A (e.g.
"Monday 1st June 2026"); the rows beneath a header are that day's collections.
Each load carries a driver, arrival/depart times, volume, temperature and a
sample number that matches the NML milk-quality sample for the same load.

Temperatures may be a single value (``4.2``) or a per-fill list
(``2.8/4.2`` or ``4.3/5.0/4.1``); split fills are averaged (unweighted).
"""

from __future__ import annotations

import datetime as dt
import io
import re
from typing import Any

import openpyxl

# 1-based spreadsheet columns -> 0-based tuple indices from iter_rows.
_COL_CUSTOMER = 0
_COL_DRIVER = 9
_COL_VEHICLE = 10
# Actual collected volume (the figure the haulier's "Total collected today"
# sums). Column 20 is the tanker's nominal capacity and overstates short loads.
_COL_VOLUME = 13
_COL_ARRIVAL = 16
_COL_DEPART = 17
_COL_TEMP = 21
_COL_SAMPLE = 22

_DATE_RE = re.compile(
    r"\b(Mon|Tues|Wednes|Thurs|Fri|Satur|Sun)day\s+"
    r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+)\s+(\d{2,})",
    re.IGNORECASE,
)

_WEEKDAYS = {
    "monday": 0,
    "tuesday": 1,
    "wednesday": 2,
    "thursday": 3,
    "friday": 4,
    "saturday": 5,
    "sunday": 6,
}

# Map the haulier's customer/origin label to our farm code.
_FARM_BY_CUSTOMER = (
    ("cwrt malle", "CM"),
    ("green acre", "GAD"),
)


def _year_candidates(year_str: str, prev_date: dt.date | None) -> list[int]:
    """Plausible 4-digit years from a (possibly fat-fingered) digit run.

    e.g. '2026' -> [2026]; '26' -> [2026] (using the previous header's century);
    '20226' -> [2022, 2026, ...]. Used with the header weekday to recover typos
    like 'June 20226' or truncated years like 'July 26'.
    """
    candidates: list[int] = []
    seen: set[int] = set()

    def add(value: int) -> None:
        if value not in seen and 1990 <= value <= 2100:
            seen.add(value)
            candidates.append(value)

    if len(year_str) == 2:
        yy = int(year_str)
        century = (prev_date.year // 100) * 100 if prev_date else 2000
        add(century + yy)
        # Also try neighbouring centuries when the previous header is missing.
        add(2000 + yy)
        add(1900 + yy)
        return candidates

    if len(year_str) == 4:
        add(int(year_str))
        return candidates

    options = [year_str[:4], year_str[-4:]]
    for i in range(len(year_str)):
        trimmed = year_str[:i] + year_str[i + 1 :]
        if len(trimmed) == 4:
            options.append(trimmed)
    for opt in options:
        add(int(opt))
    return candidates


def _parse_date_header(value: Any, prev_date: dt.date | None) -> dt.date | None:
    if not isinstance(value, str):
        return None
    match = _DATE_RE.search(value)
    if not match:
        return None
    weekday_prefix, day, month_name, year_str = match.groups()
    try:
        month = dt.datetime.strptime(month_name[:3], "%b").month
        day_num = int(day)
    except ValueError:
        return None

    expected_weekday = _WEEKDAYS.get((weekday_prefix + "day").lower())
    candidates: list[dt.date] = []
    for year in _year_candidates(year_str, prev_date):
        try:
            candidates.append(dt.date(year, month, day_num))
        except ValueError:
            continue
    if not candidates:
        return None

    # Prefer a date whose weekday matches the header; break ties by closeness to
    # the previous accepted date (these reports are sequential within a month).
    def score(d: dt.date) -> tuple[int, int]:
        weekday_ok = 0 if (expected_weekday is None or d.weekday() == expected_weekday) else 1
        distance = abs((d - prev_date).days) if prev_date else 0
        return (weekday_ok, distance)

    return min(candidates, key=score)


def _farm_for_customer(value: Any) -> str | None:
    text = str(value or "").strip().lower()
    for needle, farm in _FARM_BY_CUSTOMER:
        if needle in text:
            return farm
    return None


def _avg_temp(value: Any) -> tuple[float | None, str | None]:
    """Return (average temp, raw text). Split fills like '2.8/4.2' are averaged."""
    if value is None:
        return (None, None)
    if isinstance(value, (int, float)):
        return (float(value), None)
    raw = str(value).strip()
    if not raw:
        return (None, None)
    parts = [p.strip() for p in raw.split("/") if p.strip()]
    nums: list[float] = []
    for part in parts:
        try:
            nums.append(float(part))
        except ValueError:
            continue
    if not nums:
        return (None, raw)
    avg = round(sum(nums) / len(nums), 2)
    # Only keep raw when it actually encodes multiple fills.
    return (avg, raw if len(nums) > 1 else None)


def _norm_sample(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _as_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return None


def _as_time(value: Any) -> dt.time | None:
    if isinstance(value, dt.time):
        return value
    if isinstance(value, dt.datetime):
        return value.time()
    return None


def parse_haulier_xlsx(content: bytes) -> dict[str, Any]:
    """Parse haulier XLSX bytes into per-load collection records.

    Returns {"rows": [ {farm, collection_date, sample_id, driver, vehicle_reg,
    arrival_time, depart_time, volume_litres, temp_c, temp_raw}, ... ]}.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: list[dict[str, Any]] = []
    current_date: dt.date | None = None

    for raw_row in ws.iter_rows(values_only=True):
        if not raw_row:
            continue

        def cell(idx: int) -> Any:
            return raw_row[idx] if idx < len(raw_row) else None

        col_a = cell(_COL_CUSTOMER)

        header_date = _parse_date_header(col_a, current_date)
        if header_date is not None:
            current_date = header_date
            continue

        sample_id = _norm_sample(cell(_COL_SAMPLE))
        volume = _as_int(cell(_COL_VOLUME))
        arrival = _as_time(cell(_COL_ARRIVAL))

        if current_date is None:
            continue
        # A sampled load needs a volume or arrival time. The haulier occasionally
        # leaves the sample number blank; still keep it if it has both a volume
        # and an arrival time (enough to tell a real load from a summary row).
        if sample_id:
            if volume is None and arrival is None:
                continue
        elif volume is None or arrival is None:
            continue

        farm = _farm_for_customer(col_a)
        temp_c, temp_raw = _avg_temp(cell(_COL_TEMP))
        driver = str(cell(_COL_DRIVER) or "").strip() or None
        vehicle = str(cell(_COL_VEHICLE) or "").strip() or None

        rows.append(
            {
                "farm": farm,
                "collection_date": current_date,
                "sample_id": sample_id,
                "driver": driver,
                "vehicle_reg": vehicle,
                "arrival_time": arrival,
                "depart_time": _as_time(cell(_COL_DEPART)),
                "volume_litres": volume,
                "temp_c": temp_c,
                "temp_raw": temp_raw,
            }
        )

    wb.close()
    return {"rows": rows}
