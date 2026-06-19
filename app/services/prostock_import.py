"""Parse Prostock veterinary Excel exports (amount + quantity sheets)."""

from __future__ import annotations

import datetime
import io
from typing import Any

from openpyxl import load_workbook


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _float_or_none(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _parse_month_header(value) -> datetime.date | None:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return datetime.date(value.year, value.month, 1)
    if isinstance(value, datetime.date):
        return datetime.date(value.year, value.month, 1)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m", "%Y-%m-%d", "%m/%Y", "%b-%y", "%b-%Y"):
        try:
            parsed = datetime.datetime.strptime(text, fmt)
            return datetime.date(parsed.year, parsed.month, 1)
        except ValueError:
            continue
    return None


def _sheet_to_long(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header = rows[0]
    month_columns: list[tuple[int, datetime.date]] = []
    for col_idx in range(1, len(header)):
        month = _parse_month_header(header[col_idx])
        if month is not None:
            month_columns.append((col_idx, month))

    long_rows: list[dict[str, Any]] = []
    for row in rows[2:]:
        if not row:
            continue
        drug_name = _cell_str(row[0] if len(row) > 0 else None)
        if not drug_name:
            continue
        for col_idx, month in month_columns:
            value = row[col_idx] if col_idx < len(row) else None
            numeric = _float_or_none(value)
            if numeric is None:
                continue
            long_rows.append(
                {
                    "drug_name": drug_name,
                    "invoice_date": month,
                    "value": numeric,
                }
            )
    return long_rows


def _merge_amount_quantity(
    amount_rows: list[dict[str, Any]],
    quantity_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    amount_map: dict[tuple[str, datetime.date], float] = {}
    quantity_map: dict[tuple[str, datetime.date], float] = {}

    for row in amount_rows:
        key = (row["drug_name"], row["invoice_date"])
        amount_map[key] = row["value"]

    for row in quantity_rows:
        key = (row["drug_name"], row["invoice_date"])
        quantity_map[key] = row["value"]

    keys = set(amount_map) | set(quantity_map)
    merged: list[dict[str, Any]] = []
    for drug_name, invoice_date in sorted(keys, key=lambda k: (k[1], k[0])):
        merged.append(
            {
                "product_description": drug_name,
                "invoice_date": invoice_date,
                "goods_value": amount_map.get((drug_name, invoice_date)),
                "quantity": quantity_map.get((drug_name, invoice_date)),
            }
        )
    return merged


def parse_prostock_file(file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse Prostock workbook sheets 'amount' and 'quantity' into row dicts."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    try:
        if "amount" not in wb.sheetnames or "quantity" not in wb.sheetnames:
            names = ", ".join(wb.sheetnames)
            raise ValueError(f"Workbook must contain 'amount' and 'quantity' sheets. Found: {names}")

        amount_rows = _sheet_to_long(wb["amount"])
        quantity_rows = _sheet_to_long(wb["quantity"])
        if not amount_rows and not quantity_rows:
            return []

        return _merge_amount_quantity(amount_rows, quantity_rows)
    finally:
        wb.close()
