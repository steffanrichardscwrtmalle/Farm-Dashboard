"""Query milk collections, joined to NML quality, for the dashboard page.

Each collection (one tanker load) is matched to its NML milk-quality sample by
normalised sample number and collection date (allowing +/- 1 day, since the lab
and haulier can disagree by a day).
"""

from __future__ import annotations

import csv
import datetime as dt
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models import HERD_FARM_OPTIONS, MilkCollection, NmlMilkResult
from app.services.cows_in_milk import cows_in_milk_for_dates
from app.services.events_common import normalize_farms

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Provenance markers for user/seed loads (kept through email month-dedupe).
MANUAL_SOURCE_FILE = "manual"
SEED_GAD_MILK_SOURCE_FILE = "seed:gadmilk.xlsx"
EDITABLE_COLLECTION_SOURCES = frozenset(
    {MANUAL_SOURCE_FILE, SEED_GAD_MILK_SOURCE_FILE}
)

# Stable synthetic arrival times so Load 1/2/3 stay unique without sample IDs.
_MANUAL_LOAD_ARRIVALS = (
    dt.time(1, 0),
    dt.time(2, 0),
    dt.time(3, 0),
)


def is_editable_collection_source(source_file: str | None) -> bool:
    return (source_file or "") in EDITABLE_COLLECTION_SOURCES

_HEADERS = (
    "Farm",
    "Date",
    "Sample",
    "Driver",
    "Vehicle",
    "Arrival",
    "Depart",
    "Volume (L)",
    "Temp (°C)",
    "Temp detail",
    "Cows in milk",
    "Butterfat %",
    "Protein %",
    "SCC",
    "BactoScan",
    "Urea %",
    "A/B",
)

# Quality fields copied onto a matched collection row.
_NML_FIELDS = ("butterfat_pct", "protein_pct", "scc", "bactoscan", "urea_pct")


def _norm_sample(value: str | None) -> str:
    s = (value or "").strip().lstrip("0")
    return s or "0"


def _ab_label(value: bool | None) -> str:
    if value is True:
        return "Pass"
    if value is False:
        return "Fail"
    return ""


def _time_str(value: dt.time | None) -> str:
    return value.strftime("%H:%M") if value else ""


def _build_nml_index(
    rows: list[NmlMilkResult],
) -> dict[tuple[str, str], list[NmlMilkResult]]:
    index: dict[tuple[str, str], list[NmlMilkResult]] = {}
    for row in rows:
        if not row.farm or not row.sample_date:
            continue
        index.setdefault((row.farm, _norm_sample(row.sample_id)), []).append(row)
    return index


def _match_nml(
    collection: MilkCollection,
    index: dict[tuple[str, str], list[NmlMilkResult]],
) -> NmlMilkResult | None:
    candidates = index.get((collection.farm or "", _norm_sample(collection.sample_id)))
    if not candidates:
        return None
    best: NmlMilkResult | None = None
    best_distance = 2  # only accept +/- 1 day
    for cand in candidates:
        distance = abs((cand.sample_date - collection.collection_date).days)
        if distance < best_distance:
            best = cand
            best_distance = distance
    return best


def _row_to_dict(
    collection: MilkCollection, nml: NmlMilkResult | None
) -> dict[str, Any]:
    row: dict[str, Any] = {
        "id": collection.id,
        "farm": collection.farm or "",
        "collection_date": collection.collection_date.isoformat()
        if collection.collection_date
        else "",
        "sample_id": collection.sample_id or "",
        "driver": collection.driver or "",
        "vehicle_reg": collection.vehicle_reg or "",
        "arrival_time": _time_str(collection.arrival_time),
        "depart_time": _time_str(collection.depart_time),
        "volume_litres": collection.volume_litres,
        "temp_c": collection.temp_c,
        "temp_raw": collection.temp_raw or "",
        "cows_in_milk": collection.cows_in_milk,
        "matched": nml is not None,
        "antibiotic_pass": nml.antibiotic_pass if nml else None,
        "manual": is_editable_collection_source(collection.source_file),
    }
    for field in _NML_FIELDS:
        row[field] = getattr(nml, field) if nml else None
    return row


def _summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    def avg(metric: str, dp: int = 2) -> float | None:
        values = [float(r[metric]) for r in rows if r.get(metric) is not None]
        return round(sum(values) / len(values), dp) if values else None

    volumes = [r["volume_litres"] for r in rows if r.get("volume_litres") is not None]
    latest = max((r["collection_date"] for r in rows if r["collection_date"]), default="")
    total_volume = sum(volumes) if volumes else 0
    collection_days = {
        r["collection_date"]
        for r in rows
        if r.get("collection_date") and r.get("volume_litres") is not None
    }
    avg_daily_volume = round(total_volume / len(collection_days)) if collection_days else None
    by_day: dict[tuple[str, str], dict[str, float | None]] = {}
    for row in rows:
        date = row.get("collection_date")
        volume = row.get("volume_litres")
        if not date or volume is None:
            continue
        key = (row.get("farm") or "", str(date))
        bucket = by_day.setdefault(key, {"vol": 0.0, "cows": None})
        bucket["vol"] = float(bucket["vol"] or 0.0) + float(volume)
        cows = row.get("cows_in_milk")
        if cows is not None:
            try:
                bucket["cows"] = float(cows)
            except (TypeError, ValueError):
                pass
    daily_per_cow = [
        float(bucket["vol"]) / float(bucket["cows"])
        for bucket in by_day.values()
        if bucket.get("cows") and float(bucket["cows"]) > 0
    ]
    avg_litres_per_cow = (
        round(sum(daily_per_cow) / len(daily_per_cow), 1) if daily_per_cow else None
    )
    return {
        "count": len(rows),
        "latest_collection_date": latest,
        "total_volume": total_volume,
        "avg_volume": round(sum(volumes) / len(volumes)) if volumes else None,
        "avg_daily_volume": avg_daily_volume,
        "avg_litres_per_cow": avg_litres_per_cow,
        "days": len(collection_days),
        "avg_temp": avg("temp_c", 2),
        "matched_count": sum(1 for r in rows if r.get("matched")),
        "unmatched_count": sum(1 for r in rows if not r.get("matched")),
        "avg_butterfat_pct": avg("butterfat_pct"),
        "avg_protein_pct": avg("protein_pct"),
    }


# Per-day trend metrics: volume is summed, the rest averaged.
_TREND_SUM = ("volume_litres",)
_TREND_AVG = ("temp_c", "butterfat_pct", "protein_pct", "scc", "bactoscan", "urea_pct")


def _build_trend(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    buckets: dict[tuple[str, str], dict[str, list[float]]] = {}
    for row in rows:
        farm = row["farm"] or "?"
        date = row["collection_date"]
        if not date:
            continue
        bucket = buckets.setdefault(
            (farm, date),
            {m: [] for m in (*_TREND_SUM, *_TREND_AVG, "cows_in_milk")},
        )
        for metric in (*_TREND_SUM, *_TREND_AVG):
            value = row.get(metric)
            if value is not None:
                bucket[metric].append(float(value))
        cows = row.get("cows_in_milk")
        if cows is not None:
            try:
                cows_f = float(cows)
            except (TypeError, ValueError):
                cows_f = None
            if cows_f is not None and cows_f > 0:
                bucket["cows_in_milk"].append(cows_f)

    trend: dict[str, list[dict[str, Any]]] = {}
    for (farm, date), metrics in buckets.items():
        point: dict[str, Any] = {"date": date}
        for metric in _TREND_SUM:
            values = metrics[metric]
            point[metric] = round(sum(values)) if values else None
        for metric in _TREND_AVG:
            values = metrics[metric]
            point[metric] = round(sum(values) / len(values), 3) if values else None
        cows_values = metrics["cows_in_milk"]
        cows = max(cows_values) if cows_values else None
        volume = point.get("volume_litres")
        if volume is not None and cows:
            point["litres_per_cow"] = round(volume / cows, 2)
        else:
            point["litres_per_cow"] = None
        trend.setdefault(farm, []).append(point)
    for points in trend.values():
        points.sort(key=lambda item: item["date"])
    return trend


def list_collections(
    db: Session,
    *,
    farms: list[str] | None = None,
    date_from: dt.date | None = None,
    date_to: dt.date | None = None,
) -> dict[str, Any]:
    selected_farms = normalize_farms(farms)
    if not selected_farms:
        return {"rows": [], "total": 0, "summary": _summary([]), "trend": {}}

    query = select(MilkCollection).where(MilkCollection.farm.in_(selected_farms))
    if date_from is not None:
        query = query.where(MilkCollection.collection_date >= date_from)
    if date_to is not None:
        query = query.where(MilkCollection.collection_date <= date_to)
    query = query.order_by(
        MilkCollection.collection_date.desc(), MilkCollection.sample_id.desc()
    )
    collections = db.scalars(query).all()

    # Pull NML rows for the same farms within a +/- 1 day buffer for matching.
    nml_query = select(NmlMilkResult).where(NmlMilkResult.farm.in_(selected_farms))
    if date_from is not None:
        nml_query = nml_query.where(
            NmlMilkResult.sample_date >= date_from - dt.timedelta(days=1)
        )
    if date_to is not None:
        nml_query = nml_query.where(
            NmlMilkResult.sample_date <= date_to + dt.timedelta(days=1)
        )
    index = _build_nml_index(list(db.scalars(nml_query).all()))

    rows = [_row_to_dict(c, _match_nml(c, index)) for c in collections]
    _attach_cows_in_milk(db, rows)
    return {
        "rows": rows,
        "total": len(rows),
        "summary": _summary(rows),
        "trend": _build_trend(rows),
    }


def _attach_cows_in_milk(db: Session, rows: list[dict[str, Any]]) -> None:
    """Overwrite cows_in_milk from inventory/events (CM + GAD)."""
    if not rows:
        return
    farms: set[str] = set()
    dates: set[dt.date] = set()
    parsed: list[tuple[dict[str, Any], str, dt.date | None]] = []
    for row in rows:
        farm = (row.get("farm") or "").strip().upper()
        raw_date = row.get("collection_date") or ""
        day: dt.date | None
        if isinstance(raw_date, dt.date):
            day = raw_date
        else:
            try:
                day = dt.date.fromisoformat(str(raw_date)[:10])
            except ValueError:
                day = None
        parsed.append((row, farm, day))
        if farm and day is not None:
            farms.add(farm)
            dates.add(day)
    counts = cows_in_milk_for_dates(db, farms, dates)
    for row, farm, day in parsed:
        if not farm or day is None:
            row["cows_in_milk"] = None
            continue
        row["cows_in_milk"] = counts.get((farm, day))


def get_manual_collection_day(
    db: Session,
    *,
    farm: str,
    collection_date: dt.date,
) -> dict[str, Any]:
    """Return manual loads for one farm/day (for the edit form)."""
    farm_key = (farm or "").strip().upper()
    if farm_key not in HERD_FARM_OPTIONS:
        raise ValueError("farm must be CM or GAD")
    rows = db.scalars(
        select(MilkCollection)
        .where(
            MilkCollection.farm == farm_key,
            MilkCollection.collection_date == collection_date,
            MilkCollection.source_file.in_(EDITABLE_COLLECTION_SOURCES),
        )
        .order_by(MilkCollection.arrival_time.asc(), MilkCollection.id.asc())
    ).all()
    if not rows:
        raise ValueError("No manual collection found for that farm and date")
    cows = next((r.cows_in_milk for r in rows if r.cows_in_milk is not None), None)
    loads = [
        {
            "volume_litres": row.volume_litres,
            "temp_c": row.temp_c,
            "sample_id": row.sample_id or "",
        }
        for row in rows[:3]
    ]
    while len(loads) < 3:
        loads.append({"volume_litres": None, "temp_c": None, "sample_id": ""})
    return {
        "farm": farm_key,
        "collection_date": collection_date.isoformat(),
        "cows_in_milk": cows,
        "loads": loads,
    }


def delete_manual_collection_day(
    db: Session,
    *,
    farm: str,
    collection_date: dt.date,
) -> dict[str, Any]:
    """Delete editable (manual/seed) loads for one farm/day."""
    farm_key = (farm or "").strip().upper()
    if farm_key not in HERD_FARM_OPTIONS:
        raise ValueError("farm must be CM or GAD")
    result = db.execute(
        delete(MilkCollection).where(
            MilkCollection.farm == farm_key,
            MilkCollection.collection_date == collection_date,
            MilkCollection.source_file.in_(EDITABLE_COLLECTION_SOURCES),
        )
    )
    deleted = int(result.rowcount or 0)
    if deleted <= 0:
        raise ValueError("No manual collection found for that farm and date")
    db.commit()
    return {
        "farm": farm_key,
        "collection_date": collection_date.isoformat(),
        "loads_deleted": deleted,
    }


def create_manual_collection(
    db: Session,
    *,
    collection_date: dt.date,
    farm: str,
    loads: list[dict[str, Any]],
    cows_in_milk: int | None = None,
    replace_farm: str | None = None,
    replace_date: dt.date | None = None,
) -> dict[str, Any]:
    """Insert manual Load 1–3 rows for one farm/day (replaces prior manuals that day)."""
    farm_key = (farm or "").strip().upper()
    if farm_key not in HERD_FARM_OPTIONS:
        raise ValueError("farm must be CM or GAD")
    if cows_in_milk is not None and cows_in_milk < 0:
        raise ValueError("cows_in_milk must be zero or greater")

    cleaned_loads: list[dict[str, Any]] = []
    seen_samples: set[str] = set()
    for raw in (loads or [])[:3]:
        volume = raw.get("volume_litres")
        temp = raw.get("temp_c")
        sample_raw = str(raw.get("sample_id") or "").strip()
        sample_id = sample_raw or None
        if volume is None and temp is None and not sample_id:
            continue
        vol_int: int | None = None
        if volume is not None and volume != "":
            try:
                vol_int = int(round(float(volume)))
            except (TypeError, ValueError) as exc:
                raise ValueError("volume_litres must be a number") from exc
            if vol_int < 0:
                raise ValueError("volume_litres must be zero or greater")
        temp_f: float | None = None
        if temp is not None and temp != "":
            try:
                temp_f = round(float(temp), 2)
            except (TypeError, ValueError) as exc:
                raise ValueError("temp_c must be a number") from exc
        if sample_id:
            key = sample_id.lstrip("0") or "0"
            if key in seen_samples:
                raise ValueError(f"Duplicate sample number: {sample_id}")
            seen_samples.add(key)
        cleaned_loads.append(
            {
                "volume_litres": vol_int,
                "temp_c": temp_f,
                "sample_id": sample_id,
            }
        )

    if not cleaned_loads:
        raise ValueError("Enter at least one load volume, temperature, or sample")

    # Clear previous manuals for this save target and, when editing, the original day.
    clear_keys = {(farm_key, collection_date)}
    if replace_farm and replace_date:
        replace_key = (replace_farm.strip().upper(), replace_date)
        if replace_key[0] in HERD_FARM_OPTIONS:
            clear_keys.add(replace_key)
    for clear_farm, clear_date in clear_keys:
        db.execute(
            delete(MilkCollection).where(
                MilkCollection.farm == clear_farm,
                MilkCollection.collection_date == clear_date,
                MilkCollection.source_file.in_(EDITABLE_COLLECTION_SOURCES),
            )
        )

    for load in cleaned_loads:
        sample_id = load["sample_id"]
        if not sample_id:
            continue
        clash = db.scalar(
            select(MilkCollection.id).where(
                MilkCollection.farm == farm_key,
                MilkCollection.collection_date == collection_date,
                MilkCollection.sample_id == sample_id,
            )
        )
        if clash is not None:
            raise ValueError(
                f"Sample {sample_id} already exists for {farm_key} on "
                f"{collection_date.isoformat()}"
            )

    now = dt.datetime.now(dt.timezone.utc).replace(tzinfo=None)
    created: list[dict[str, Any]] = []
    for index, load in enumerate(cleaned_loads):
        arrival = _MANUAL_LOAD_ARRIVALS[index]
        row = MilkCollection(
            farm=farm_key,
            collection_date=collection_date,
            sample_id=load["sample_id"],
            driver=None,
            vehicle_reg=None,
            arrival_time=arrival,
            depart_time=None,
            volume_litres=load["volume_litres"],
            temp_c=load["temp_c"],
            temp_raw=None,
            cows_in_milk=cows_in_milk,
            source_message_id=f"manual:{farm_key}:{collection_date.isoformat()}",
            source_file=MANUAL_SOURCE_FILE,
            source_received=now,
        )
        db.add(row)
        created.append(
            {
                "farm": farm_key,
                "collection_date": collection_date.isoformat(),
                "load": index + 1,
                "volume_litres": load["volume_litres"],
                "temp_c": load["temp_c"],
                "sample_id": load["sample_id"] or "",
                "cows_in_milk": cows_in_milk,
            }
        )
    db.commit()
    return {
        "farm": farm_key,
        "collection_date": collection_date.isoformat(),
        "cows_in_milk": cows_in_milk,
        "loads_created": len(created),
        "rows": created,
    }


def _export_cells(row: dict[str, Any]) -> list[Any]:
    return [
        row.get("farm", ""),
        row.get("collection_date", ""),
        row.get("sample_id", ""),
        row.get("driver", ""),
        row.get("vehicle_reg", ""),
        row.get("arrival_time", ""),
        row.get("depart_time", ""),
        row.get("volume_litres"),
        row.get("temp_c"),
        row.get("temp_raw", ""),
        row.get("cows_in_milk"),
        row.get("butterfat_pct"),
        row.get("protein_pct"),
        row.get("scc"),
        row.get("bactoscan"),
        row.get("urea_pct"),
        _ab_label(row.get("antibiotic_pass")),
    ]


def build_collections_csv(rows: list[dict[str, Any]]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(_HEADERS)
    for row in rows:
        writer.writerow(_export_cells(row))
    return buffer.getvalue()


def build_collections_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Milk Collections"
    ws.append(list(_HEADERS))
    for row in rows:
        ws.append(_export_cells(row))

    widths = [8, 12, 10, 16, 12, 9, 9, 11, 10, 14, 12, 11, 10, 8, 11, 8, 7]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
