"""Farm Reports: inventory-based animal lists with widget counts."""

from __future__ import annotations

import datetime as dt
import io
import math
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import Float, Integer, and_, cast, func, or_, select
from sqlalchemy.orm import Session

from app.models import HerdInventory
from app.services.farm_schedule import FARM_LABELS, normalize_farm

HEIFERS_TO_SCAN_RC = (3, 4)
HEIFERS_TO_SCAN_MIN_DSLH = 31
COLLARS_TO_PUT_ON_RC = 0
COLLARS_BROKEN_RC = (0, 3, 4)
COLLARS_TO_PUT_ON_MIN_EWGT = 385
BROKEN_COLLAR_FILL = "FECACA"
BROKEN_COLLAR_TEXT = "7F1D1D"
WIDGET_HEIFERS_TO_SCAN = "heifers-to-scan"
WIDGET_COLLARS_TO_PUT_ON = "collars-to-put-on"
WIDGET_HEIFERS_TO_SCAN_AND_COLLARS = "heifers-to-scan-and-collars"
BLANK_PEN = "__blank__"
NO_MATCH_PEN = "__no_match__"
PDF_CONTENT_TYPE = "application/pdf"
XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_ROWS_PER_PAGE = 40
PDF_TABLE_FONT_SIZE = 10
PDF_COMPACT_FONT_SIZE = 7
PDF_COMPACT_COL_KEYS = frozenset({"reason", "httag"})
ETAG4_COL_WIDTH_MM = 16.0
ETAG4_XLSX_COL_WIDTH = 9
HEIFERS_TO_SCAN_COLUMNS: tuple[tuple[str, str], ...] = (
    ("id", "ID"),
    ("remark", "REMARK"),
    ("etag4", "ETAG4"),
    ("dslh", "DSLH"),
    ("tbrd", "TBRD"),
    ("rpro", "RPRO"),
    ("pen", "PEN"),
)
COLLARS_TO_PUT_ON_COLUMNS: tuple[tuple[str, str], ...] = (
    ("id", "ID"),
    ("remark", "REMARK"),
    ("etag4", "ETAG4"),
    ("ewgt", "EWGT"),
    ("httag", "HTTAG"),
    ("aged", "AGED"),
    ("rpro", "RPRO"),
    ("pen", "PEN"),
)
REASON_PREG_CHECK = "Preg Check"
REASON_FAULTY_COLLAR = "Faulty Collar"
REASON_PUT_COLLAR_ON = "Put Collar On"
REASON_PREG_CHECK_AND_FAULTY = "Preg Check & Faulty Collar"
HEIFERS_TO_SCAN_AND_COLLARS_COLUMNS: tuple[tuple[str, str], ...] = (
    ("id", "ID"),
    ("remark", "REMARK"),
    ("reason", "REASON"),
    ("etag4", "ETAG4"),
    ("dslh", "DSLH"),
    ("tbrd", "TBRD"),
    ("ewgt", "EWGT"),
    ("httag", "HTTAG"),
    ("aged", "AGED"),
    ("rpro", "RPRO"),
    ("pen", "PEN"),
)
REPORT_SPECS: dict[str, dict[str, Any]] = {
    WIDGET_HEIFERS_TO_SCAN: {
        "title": "Heifers To Scan",
        "filename": "heifers_to_scan",
        "columns": HEIFERS_TO_SCAN_COLUMNS,
        "xlsx_widths": [10, 24, 10, 8, 8, 10, 8],
        "pdf_raw_mm": (20, 72, 22, 18, 16, 24, 22),
    },
    WIDGET_COLLARS_TO_PUT_ON: {
        "title": "Collars To Put On",
        "filename": "collars_to_put_on",
        "columns": COLLARS_TO_PUT_ON_COLUMNS,
        "xlsx_widths": [10, 20, 10, 8, 8, 8, 10, 8],
        "pdf_raw_mm": (18, 52, 20, 18, 18, 16, 22, 16),
    },
    WIDGET_HEIFERS_TO_SCAN_AND_COLLARS: {
        "title": "Heifers To Scan & Collars",
        "filename": "heifers_to_scan_and_collars",
        "columns": HEIFERS_TO_SCAN_AND_COLLARS_COLUMNS,
        "xlsx_widths": [10, 18, 22, 10, 8, 8, 8, 8, 8, 10, 8],
        "pdf_raw_mm": (14, 36, 34, 16, 12, 10, 14, 12, 12, 16, 14),
    },
}


def etag4(etag: str | None) -> str | None:
    """Last four digits of ETAG, keeping a leading zero when present."""
    trimmed = (etag or "").strip()
    digits = "".join(ch for ch in trimmed if ch.isdigit())
    return digits[-4:] if digits else None


def _whole_number(value: object) -> int | float | None:
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number.is_integer():
        return int(number)
    return number


def _cell(value: object) -> str:
    if value is None or value == "":
        return ""
    return str(value)


def _etag4_sort_key(value: object) -> tuple[int, int, str]:
    text = str(value or "").strip()
    if text.isdigit():
        return (0, int(text), "")
    if not text:
        return (2, 0, "")
    return (1, 0, text.lower())


def sort_heifers_by_etag4(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda row: _etag4_sort_key(row.get("etag4")))


def etag4_column_grid(
    rows: list[dict[str, Any]], *, rows_per_page: int = PDF_ROWS_PER_PAGE
) -> list[list[str]]:
    """Lay ETAG4 values out in column-major order, 40 rows per column."""
    tags = [_cell(row.get("etag4")) for row in rows]
    if not tags:
        return []
    columns = max(1, math.ceil(len(tags) / rows_per_page))
    grid = [[""] * columns for _ in range(rows_per_page)]
    for index, tag in enumerate(tags):
        grid[index % rows_per_page][index // rows_per_page] = tag
    while grid and not any(grid[-1]):
        grid.pop()
    return grid


def etag4_flag_grid(
    rows: list[dict[str, Any]], *, rows_per_page: int = PDF_ROWS_PER_PAGE
) -> list[list[bool]]:
    flags = [bool(row.get("broken_collar")) for row in rows]
    if not flags:
        return []
    columns = max(1, math.ceil(len(flags) / rows_per_page))
    grid = [[False] * columns for _ in range(rows_per_page)]
    for index, flag in enumerate(flags):
        grid[index % rows_per_page][index // rows_per_page] = flag
    values = etag4_column_grid(rows, rows_per_page=rows_per_page)
    return grid[: len(values)]


def etag4_pdf_col_widths(columns: int, usable_w: float) -> list[float]:
    """Compact ETAG4 columns; shrink only if they would overflow the page."""
    from reportlab.lib.units import mm

    preferred = ETAG4_COL_WIDTH_MM * mm
    total = preferred * max(columns, 1)
    width = preferred if total <= usable_w else usable_w / max(columns, 1)
    return [width] * max(columns, 1)


def _pen_sort_key(pen: str) -> tuple[int, int, str]:
    if pen.isdigit():
        return (0, int(pen), "")
    return (1, 0, pen.lower())


def unique_pens(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    seen: set[str] = set()
    pens: list[str] = []
    has_blank = False
    for row in rows:
        pen = (row.get("pen") or "").strip()
        if not pen:
            has_blank = True
            continue
        if pen not in seen:
            seen.add(pen)
            pens.append(pen)
    pens.sort(key=_pen_sort_key)
    items = [{"id": pen, "label": pen} for pen in pens]
    if has_blank:
        items.append({"id": BLANK_PEN, "label": "(blank)"})
    return items


def apply_pen_filter(
    rows: list[dict[str, Any]], pens: list[str] | None
) -> list[dict[str, Any]]:
    if not pens:
        return list(rows)
    if pens == [NO_MATCH_PEN]:
        return []
    selected = {pen for pen in pens if pen and pen != NO_MATCH_PEN}
    include_blank = BLANK_PEN in selected
    selected.discard(BLANK_PEN)
    out: list[dict[str, Any]] = []
    for row in rows:
        pen = (row.get("pen") or "").strip()
        if not pen:
            if include_blank:
                out.append(row)
        elif pen in selected:
            out.append(row)
    return out


def _heifers_to_scan_query(farm_key: str):
    return (
        select(HerdInventory)
        .where(HerdInventory.farm == farm_key)
        .where(HerdInventory.category == "Youngstock")
        .where(cast(HerdInventory.rc, Integer).in_(HEIFERS_TO_SCAN_RC))
        .where(HerdInventory.dslh > HEIFERS_TO_SCAN_MIN_DSLH)
        .order_by(HerdInventory.cow_id.asc())
    )


def _httag_numeric_expr():
    return func.coalesce(
        cast(func.nullif(func.trim(HerdInventory.httag), ""), Float),
        0,
    )


def _httag_number(value: object) -> float:
    text = str(value or "").strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _is_broken_collar(*, httag: object, rum: object) -> bool:
    return _httag_number(httag) > 0 and rum is not None and float(rum) == 0


def _collars_to_put_on_query(farm_key: str):
    httag_num = _httag_numeric_expr()
    rc_int = cast(HerdInventory.rc, Integer)
    needs_collar = and_(httag_num == 0, rc_int == COLLARS_TO_PUT_ON_RC)
    broken_collar = and_(
        httag_num > 0,
        HerdInventory.rum == 0,
        rc_int.in_(COLLARS_BROKEN_RC),
    )
    return (
        select(HerdInventory)
        .where(HerdInventory.farm == farm_key)
        .where(HerdInventory.category == "Youngstock")
        .where(cast(HerdInventory.lact, Integer) == 0)
        .where(HerdInventory.ewgt >= COLLARS_TO_PUT_ON_MIN_EWGT)
        .where(or_(needs_collar, broken_collar))
        .order_by(HerdInventory.cow_id.asc())
    )


def _serialize_inventory_animal(
    row: HerdInventory, *, mark_broken_collar: bool = False
) -> dict[str, Any]:
    return {
        "id": row.cow_id,
        "remark": (row.remark or "").strip() or None,
        "etag4": etag4(row.etag),
        "dslh": _whole_number(row.dslh),
        "tbrd": row.tbrd,
        "rpro": (row.rpro or "").strip() or None,
        "pen": (row.pen or "").strip() or None,
        "ewgt": _whole_number(row.ewgt),
        "httag": _whole_number(_httag_number(row.httag)),
        "aged": row.aged,
        "broken_collar": (
            mark_broken_collar and _is_broken_collar(httag=row.httag, rum=row.rum)
        ),
    }


def _animal_list_report(
    db: Session,
    farm: str,
    *,
    widget_id: str,
    title: str,
    query,
    pens: list[str] | None = None,
) -> dict[str, Any]:
    farm_key = normalize_farm(farm)
    rows = db.scalars(query).all()
    animals = [
        _serialize_inventory_animal(
            row,
            mark_broken_collar=widget_id
            in {WIDGET_COLLARS_TO_PUT_ON, WIDGET_HEIFERS_TO_SCAN_AND_COLLARS},
        )
        for row in rows
    ]
    filtered = sort_heifers_by_etag4(apply_pen_filter(animals, pens))
    return {
        "id": widget_id,
        "title": title,
        "count": len(animals),
        "filtered_count": len(filtered),
        "pens": unique_pens(animals),
        "rows": filtered,
        "farm": farm_key,
        "farm_label": FARM_LABELS[farm_key],
    }


def heifers_to_scan(
    db: Session, farm: str, pens: list[str] | None = None
) -> dict[str, Any]:
    farm_key = normalize_farm(farm)
    return _animal_list_report(
        db,
        farm_key,
        widget_id=WIDGET_HEIFERS_TO_SCAN,
        title=REPORT_SPECS[WIDGET_HEIFERS_TO_SCAN]["title"],
        query=_heifers_to_scan_query(farm_key),
        pens=pens,
    )


def collars_to_put_on(
    db: Session, farm: str, pens: list[str] | None = None
) -> dict[str, Any]:
    farm_key = normalize_farm(farm)
    return _animal_list_report(
        db,
        farm_key,
        widget_id=WIDGET_COLLARS_TO_PUT_ON,
        title=REPORT_SPECS[WIDGET_COLLARS_TO_PUT_ON]["title"],
        query=_collars_to_put_on_query(farm_key),
        pens=pens,
    )


def _combined_reason(
    *, preg_check: bool, faulty_collar: bool, put_collar_on: bool
) -> str:
    if preg_check and faulty_collar:
        return REASON_PREG_CHECK_AND_FAULTY
    if preg_check:
        return REASON_PREG_CHECK
    if faulty_collar:
        return REASON_FAULTY_COLLAR
    if put_collar_on:
        return REASON_PUT_COLLAR_ON
    return ""


def _merge_scan_and_collar_rows(
    scan_rows: list[dict[str, Any]], collar_rows: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    scan_by_id = {row["id"]: dict(row) for row in scan_rows}
    collar_by_id = {row["id"]: dict(row) for row in collar_rows}
    merged: list[dict[str, Any]] = []
    for cow_id in scan_by_id.keys() | collar_by_id.keys():
        collar_row = collar_by_id.get(cow_id)
        row = dict(collar_row or scan_by_id[cow_id])
        row["reason"] = _combined_reason(
            preg_check=cow_id in scan_by_id,
            faulty_collar=bool(collar_row and collar_row.get("broken_collar")),
            put_collar_on=bool(collar_row and not collar_row.get("broken_collar")),
        )
        merged.append(row)
    return merged


def _combined_scan_and_collars_report(
    scan: dict[str, Any],
    collars: dict[str, Any],
    pens: list[str] | None = None,
) -> dict[str, Any]:
    spec = REPORT_SPECS[WIDGET_HEIFERS_TO_SCAN_AND_COLLARS]
    animals = _merge_scan_and_collar_rows(scan["rows"], collars["rows"])
    filtered = sort_heifers_by_etag4(apply_pen_filter(animals, pens))
    return {
        "id": WIDGET_HEIFERS_TO_SCAN_AND_COLLARS,
        "title": spec["title"],
        "count": len(animals),
        "filtered_count": len(filtered),
        "pens": unique_pens(animals),
        "rows": filtered,
        "farm": scan["farm"],
        "farm_label": scan["farm_label"],
    }


def heifers_to_scan_and_collars(
    db: Session, farm: str, pens: list[str] | None = None
) -> dict[str, Any]:
    farm_key = normalize_farm(farm)
    return _combined_scan_and_collars_report(
        heifers_to_scan(db, farm_key),
        collars_to_put_on(db, farm_key),
        pens=pens,
    )


def load_report(
    db: Session, farm: str, report_id: str, pens: list[str] | None = None
) -> dict[str, Any]:
    loaders = {
        WIDGET_HEIFERS_TO_SCAN: heifers_to_scan,
        WIDGET_COLLARS_TO_PUT_ON: collars_to_put_on,
        WIDGET_HEIFERS_TO_SCAN_AND_COLLARS: heifers_to_scan_and_collars,
    }
    loader = loaders.get(report_id)
    if loader is None:
        raise ValueError(f"Unknown report: {report_id}")
    return loader(db, farm, pens=pens)


def farm_reports(db: Session, farm: str) -> dict[str, Any]:
    farm_key = normalize_farm(farm)
    scan = heifers_to_scan(db, farm_key)
    collars = collars_to_put_on(db, farm_key)
    combined = _combined_scan_and_collars_report(scan, collars)
    return {
        "farm": farm_key,
        "farm_label": FARM_LABELS[farm_key],
        "widgets": [
            {"id": scan["id"], "title": scan["title"], "count": scan["count"]},
            {"id": collars["id"], "title": collars["title"], "count": collars["count"]},
            {
                "id": combined["id"],
                "title": combined["title"],
                "count": combined["count"],
            },
        ],
        "heifers_to_scan": scan,
        "collars_to_put_on": collars,
        "heifers_to_scan_and_collars": combined,
    }


def _style_xlsx_header(sheet) -> None:
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.orientation = "portrait"
    sheet.print_title_rows = "1:1"


def _broken_fill():
    return PatternFill(fill_type="solid", fgColor=BROKEN_COLLAR_FILL)


def _broken_font(*, bold: bool = False) -> Font:
    return Font(bold=bold, color=BROKEN_COLLAR_TEXT)


def _write_etag4_xlsx_sheet(
    sheet, grid: list[list[str]], broken_grid: list[list[bool]] | None = None
) -> None:
    if not grid:
        sheet.append(["ETAG4"])
        _style_xlsx_header(sheet)
        sheet.column_dimensions["A"].width = ETAG4_XLSX_COL_WIDTH
        return
    columns = len(grid[0])
    sheet.append(["ETAG4"] * columns)
    fill = _broken_fill()
    font = _broken_font()
    for row_index, row in enumerate(grid):
        if not any(cell for cell in row):
            continue
        sheet.append(row)
        excel_row = sheet.max_row
        flags = broken_grid[row_index] if broken_grid and row_index < len(broken_grid) else []
        for col_index, cell in enumerate(sheet[excel_row]):
            cell.alignment = Alignment(horizontal="center")
            if col_index < len(flags) and flags[col_index]:
                cell.fill = fill
                cell.font = font
    for index in range(1, columns + 1):
        sheet.column_dimensions[get_column_letter(index)].width = ETAG4_XLSX_COL_WIDTH
    _style_xlsx_header(sheet)


def _report_spec(report: dict[str, Any]) -> dict[str, Any]:
    report_id = report.get("id") or WIDGET_HEIFERS_TO_SCAN
    spec = REPORT_SPECS.get(str(report_id))
    if spec is None:
        raise ValueError(f"Unknown report: {report_id}")
    return spec


def build_report_xlsx(report: dict[str, Any], *, etag4_only: bool = False) -> bytes:
    spec = _report_spec(report)
    columns: tuple[tuple[str, str], ...] = spec["columns"]
    wb = Workbook()
    rows = sort_heifers_by_etag4(list(report.get("rows") or []))
    grid = etag4_column_grid(rows)
    broken_grid = etag4_flag_grid(rows)

    if etag4_only:
        sheet = wb.active
        sheet.title = "ETAG4"
        _write_etag4_xlsx_sheet(sheet, grid, broken_grid)
    else:
        ws = wb.active
        ws.title = str(spec["title"])[:31]
        headers = [label for _, label in columns]
        ws.append(headers)
        fill = _broken_fill()
        font = _broken_font()
        for row in rows:
            ws.append([_cell(row.get(key)) for key, _ in columns])
            if row.get("broken_collar"):
                for cell in ws[ws.max_row]:
                    cell.fill = fill
                    cell.font = font
        for index, width in enumerate(spec["xlsx_widths"], start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
        _style_xlsx_header(ws)
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.horizontalCentered = True
        etag_sheet = wb.create_sheet("ETAG4")
        _write_etag4_xlsx_sheet(etag_sheet, grid, broken_grid)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_heifers_to_scan_xlsx(report: dict[str, Any], *, etag4_only: bool = False) -> bytes:
    return build_report_xlsx(report, etag4_only=etag4_only)


def build_report_pdf(report: dict[str, Any], *, etag4_only: bool = False) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak,
        SimpleDocTemplate,
        Table,
        TableStyle,
    )

    spec = _report_spec(report)
    columns: tuple[tuple[str, str], ...] = spec["columns"]
    report_title = str(spec["title"])
    farm_label = report.get("farm_label") or ""
    title = (
        f"{report_title} · ETAG4 · {farm_label}"
        if etag4_only
        else f"{report_title} · {farm_label}"
    )
    generated = dt.datetime.now().strftime("%d/%m/%Y %H:%M")
    count = int(report.get("filtered_count") or len(report.get("rows") or []))

    buffer = io.BytesIO()
    page_w, page_h = A4
    left = right = 8 * mm
    top = 16 * mm
    bottom = 10 * mm
    usable_w = page_w - left - right
    raw_mm = spec["pdf_raw_mm"]
    scale = usable_w / (sum(raw_mm) * mm)
    col_widths = [width * mm * scale for width in raw_mm]
    row_h = 6.2 * mm
    header_h = 6.8 * mm

    def _header(canvas, doc) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 9)
        canvas.drawString(left, page_h - 11 * mm, title)
        canvas.setFont("Helvetica", 8)
        canvas.drawString(left, page_h - 15 * mm, f"{count} animals  |  {generated}")
        canvas.drawRightString(page_w - right, page_h - 11 * mm, f"Page {doc.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        title=title,
        leftMargin=left,
        rightMargin=right,
        topMargin=top,
        bottomMargin=bottom,
    )

    header = [label for _, label in columns]
    rows = sort_heifers_by_etag4(list(report.get("rows") or []))
    data_rows = [
        [_pdf_cell(key, row.get(key)) for key, _ in columns]
        for row in rows
    ]
    elements: list[Any] = []
    empty_pad = [""] * (len(header) - 1)

    def _etag4_table(grid: list[list[str]]):
        grid_columns = len(grid[0])
        etag_widths = etag4_pdf_col_widths(grid_columns, usable_w)
        etag_header = ["ETAG4"] * grid_columns
        etag_table = Table(
            [etag_header] + grid,
            colWidths=etag_widths,
            rowHeights=[header_h] + [row_h] * len(grid),
            repeatRows=1,
            hAlign="LEFT",
        )
        etag_table.setStyle(_pdf_table_style())
        etag_table.setStyle(TableStyle([("ALIGN", (0, 1), (-1, -1), "CENTER")]))
        return etag_table

    grid = etag4_column_grid(rows)
    if etag4_only:
        if grid:
            elements.append(_etag4_table(grid))
        else:
            empty = Table(
                [["ETAG4"], ["No animals match the selected pens."]],
                colWidths=[ETAG4_COL_WIDTH_MM * mm],
                hAlign="LEFT",
            )
            empty.setStyle(_pdf_table_style())
            elements.append(empty)
    elif not data_rows:
        empty = Table(
            [header, ["No animals match the selected pens."] + empty_pad],
            colWidths=col_widths,
            repeatRows=1,
        )
        empty.setStyle(_pdf_table_style(columns))
        empty.setStyle(
            TableStyle([("SPAN", (0, 1), (-1, 1)), ("ALIGN", (0, 1), (-1, 1), "LEFT")])
        )
        elements.append(empty)
    else:
        for start in range(0, len(data_rows), PDF_ROWS_PER_PAGE):
            chunk = data_rows[start : start + PDF_ROWS_PER_PAGE]
            table = Table(
                [header] + chunk,
                colWidths=col_widths,
                rowHeights=[header_h] + [row_h] * len(chunk),
                repeatRows=1,
            )
            table.setStyle(_pdf_table_style(columns))
            elements.append(table)
            if start + PDF_ROWS_PER_PAGE < len(data_rows):
                elements.append(PageBreak())
        if grid:
            elements.append(PageBreak())
            elements.append(_etag4_table(grid))

    doc.build(elements, onFirstPage=_header, onLaterPages=_header)
    return buffer.getvalue()


def build_heifers_to_scan_pdf(report: dict[str, Any], *, etag4_only: bool = False) -> bytes:
    return build_report_pdf(report, etag4_only=etag4_only)


def _pdf_cell(key: str, value: object) -> str:
    text = _cell(value)
    if key == "remark" and len(text) > 36:
        return text[:35] + "…"
    return text


def _pdf_table_style(columns: tuple[tuple[str, str], ...] | None = None):
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    commands: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), PDF_TABLE_FONT_SIZE),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (2, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
    ]
    for index, (key, _) in enumerate(columns or ()):
        if key not in PDF_COMPACT_COL_KEYS:
            continue
        commands.append(("FONTSIZE", (index, 0), (index, -1), PDF_COMPACT_FONT_SIZE))
        commands.append(("LEFTPADDING", (index, 0), (index, -1), 1))
        commands.append(("RIGHTPADDING", (index, 0), (index, -1), 1))
    return TableStyle(commands)
