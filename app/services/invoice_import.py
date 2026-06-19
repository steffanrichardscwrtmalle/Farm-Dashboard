"""Parse Wynnstay Excel export files into row dicts."""

from __future__ import annotations

import datetime
import io
from typing import Any

from openpyxl import load_workbook

from app.constants import SOURCE_COLS_TO_COPY
from app.services.transforms import cell_to_date


def _source_to_dest_field(col_offset: int) -> str:
    if col_offset == 0:
        return "date"
    if col_offset == 1:
        return "reference"
    if col_offset == 2:
        return "product_code"
    if col_offset == 3:
        return "product_description"
    if col_offset == 4:
        return "quantity"
    if col_offset == 5:
        return "unit"
    if col_offset == 6:
        return "price"
    if col_offset == 7:
        return "goods_value"
    if col_offset == 8:
        return "vat"
    if col_offset == 9:
        return "total"
    raise ValueError(f"Unexpected column offset: {col_offset}")


def parse_import_file(
    file_bytes: bytes,
    invoice_date: datetime.date,
    date_added: datetime.date | None = None,
) -> list[dict[str, Any]]:
    """
    Parse the first sheet of a Wynnstay export (rows from A2).
    Category and farm description are left empty for mapping step.
    """
    if date_added is None:
        date_added = datetime.date.today()

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    try:
        import_ws = wb[wb.sheetnames[0]]

        last_import_row = import_ws.max_row
        while last_import_row >= 2 and import_ws.cell(row=last_import_row, column=1).value in (
            None,
            "",
        ):
            last_import_row -= 1

        if last_import_row < 2:
            return []

        rows: list[dict[str, Any]] = []
        for src_row_idx in range(2, last_import_row + 1):
            src_row = import_ws[src_row_idx]
            row: dict[str, Any] = {
                "category": "",
                "farm_description": "",
                "date_added": date_added,
                "invoice_date": invoice_date,
                "recent": "No",
            }
            for col_offset in range(SOURCE_COLS_TO_COPY):
                if col_offset < len(src_row):
                    field = _source_to_dest_field(col_offset)
                    value = src_row[col_offset].value
                    if field == "date":
                        value = cell_to_date(value) or value
                    row[field] = value
            rows.append(row)
        return rows
    finally:
        wb.close()
