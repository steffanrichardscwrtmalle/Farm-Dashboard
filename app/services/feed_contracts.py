"""Feed purchase contracts: list, create, summary, options, and Excel seed."""

from __future__ import annotations

import calendar
import datetime as dt
import io
import json
import logging
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.models import (
    FEED_PRODUCT_TYPES_DEFAULT,
    FEED_PRODUCT_TYPES_SETTING_KEY,
    FEED_PRODUCTS_SETTING_KEY,
    FEED_SUPPLIERS_SETTING_KEY,
    AppSetting,
    FeedContract,
)

logger = logging.getLogger(__name__)

_SEED_PATH = Path(__file__).resolve().parent.parent / "seed_data" / "feedcontracts.xlsx"
SEED_SOURCE_FILE = "feedcontracts.xlsx"
_OPTION_KEYS = {
    "products": FEED_PRODUCTS_SETTING_KEY,
    "product_types": FEED_PRODUCT_TYPES_SETTING_KEY,
    "suppliers": FEED_SUPPLIERS_SETTING_KEY,
}


class FeedContractError(Exception):
    """Feed contract validation or persistence failed."""


def _as_date(value: Any) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return dt.datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _clean_str(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def parse_feedcontracts_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Parse Orders sheet into contract row dicts."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        sheet_name = "Orders" if "Orders" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        rows: list[dict[str, Any]] = []
        for index, row in enumerate(ws.iter_rows(values_only=True)):
            if index == 0 or not row:
                continue
            delivery_date = _as_date(row[0] if len(row) > 0 else None)
            purchase_date = _as_date(row[1] if len(row) > 1 else None)
            product = _clean_str(row[2] if len(row) > 2 else None)
            product_type = _clean_str(row[3] if len(row) > 3 else None)
            tonnage = _as_float(row[4] if len(row) > 4 else None)
            price = _as_float(row[5] if len(row) > 5 else None)
            supplier = _clean_str(row[6] if len(row) > 6 else None)
            if not purchase_date or not delivery_date or not product or not supplier:
                continue
            if tonnage is None or price is None:
                continue
            rows.append(
                {
                    "purchase_date": purchase_date,
                    "delivery_date": delivery_date,
                    "product": product,
                    "product_type": product_type,
                    "tonnage": tonnage,
                    "price": price,
                    "supplier": supplier,
                }
            )
        return rows
    finally:
        wb.close()


def list_feed_contracts(
    db: Session,
    *,
    search: str | None = None,
    supplier: str | None = None,
    product: str | None = None,
    date_from: dt.date | None = None,
    date_to: dt.date | None = None,
) -> dict[str, Any]:
    query = select(FeedContract).order_by(
        FeedContract.delivery_date.asc(),
        FeedContract.purchase_date.asc(),
        FeedContract.id.asc(),
    )
    if supplier:
        query = query.where(FeedContract.supplier == supplier)
    if product:
        query = query.where(FeedContract.product == product)
    if date_from:
        query = query.where(FeedContract.delivery_date >= date_from)
    if date_to:
        query = query.where(FeedContract.delivery_date <= date_to)
    if search:
        term = f"%{search.strip()}%"
        query = query.where(
            or_(
                FeedContract.product.ilike(term),
                FeedContract.supplier.ilike(term),
                FeedContract.product_type.ilike(term),
            )
        )
    rows = list(db.scalars(query).all())
    options = get_feed_contract_options(db)
    return {
        "rows": [row.to_dict() for row in rows],
        "count": len(rows),
        "suppliers": options["suppliers"],
        "products": options["products"],
        "product_types": options["product_types"],
        "defaults": default_summary_date_range(db),
    }


def create_feed_contract(db: Session, payload: dict[str, Any]) -> dict[str, Any]:
    result = create_feed_contracts_bulk(
        db,
        {
            **payload,
            "delivery_months": None,
            "delivery_dates": [payload.get("delivery_date")],
        },
    )
    rows = result.get("rows") or []
    if not rows:
        raise FeedContractError("Could not create purchase.")
    return {"row": rows[0]}


def _parse_delivery_months(payload: dict[str, Any]) -> list[dt.date]:
    """Resolve delivery months/dates to first-of-month dates."""
    dates: list[dt.date] = []
    months = payload.get("delivery_months")
    if months:
        for raw in months:
            text = str(raw or "").strip()
            if not text:
                continue
            try:
                year_s, month_s = text.split("-", 1)
                year, month = int(year_s), int(month_s)
                dates.append(dt.date(year, month, 1))
            except (TypeError, ValueError):
                raise FeedContractError(f"Invalid delivery month: {text}") from None
    else:
        for raw in payload.get("delivery_dates") or []:
            day = raw if isinstance(raw, dt.date) else _as_date(raw)
            if day is None:
                continue
            dates.append(day.replace(day=1))
        single = payload.get("delivery_date")
        if single is not None:
            day = single if isinstance(single, dt.date) else _as_date(single)
            if day is not None:
                dates.append(day.replace(day=1))

    # Deduplicate while preserving order.
    seen: set[dt.date] = set()
    unique: list[dt.date] = []
    for day in dates:
        if day not in seen:
            seen.add(day)
            unique.append(day)
    return unique


def create_feed_contracts_bulk(db: Session, payload: dict[str, Any]) -> dict[str, Any]:
    """Create one contract row per delivery month, applying the same tonnage to each."""
    purchase_date = payload.get("purchase_date")
    product = _clean_str(payload.get("product"))
    product_type = _clean_str(payload.get("product_type"))
    supplier = _clean_str(payload.get("supplier"))
    tonnage = _as_float(payload.get("tonnage"))
    price = _as_float(payload.get("price"))

    if not isinstance(purchase_date, dt.date):
        purchase_date = _as_date(purchase_date)

    delivery_dates = _parse_delivery_months(payload)

    if purchase_date is None:
        raise FeedContractError("Purchase date is required.")
    if not delivery_dates:
        raise FeedContractError("Select at least one delivery month.")
    if not product:
        raise FeedContractError("Product is required.")
    if not supplier:
        raise FeedContractError("Supplier is required.")
    if tonnage is None or tonnage < 0:
        raise FeedContractError("Tonnage must be a non-negative number.")
    if price is None or price < 0:
        raise FeedContractError("Price must be a non-negative number.")

    created: list[FeedContract] = []
    for delivery_date in delivery_dates:
        row = FeedContract(
            purchase_date=purchase_date,
            delivery_date=delivery_date,
            product=product,
            product_type=product_type,
            tonnage=tonnage,
            price=price,
            supplier=supplier,
        )
        db.add(row)
        created.append(row)
    db.commit()
    for row in created:
        db.refresh(row)
    _ensure_options_include(
        db,
        product=product,
        product_type=product_type,
        supplier=supplier,
    )
    return {
        "rows": [row.to_dict() for row in created],
        "count": len(created),
        "tonnage_per_month": tonnage,
        "total_tonnage": tonnage * len(created),
    }


def delete_feed_contract(db: Session, contract_id: int) -> dict[str, Any]:
    row = db.get(FeedContract, contract_id)
    if row is None:
        raise FeedContractError("Contract not found.")
    db.delete(row)
    db.commit()
    return {"status": "deleted", "id": contract_id}


def update_feed_contract(
    db: Session,
    contract_id: int,
    payload: dict[str, Any],
) -> dict[str, Any]:
    row = db.get(FeedContract, contract_id)
    if row is None:
        raise FeedContractError("Contract not found.")

    purchase_date = payload.get("purchase_date")
    delivery_date = payload.get("delivery_date")
    product = _clean_str(payload.get("product"))
    product_type = _clean_str(payload.get("product_type"))
    supplier = _clean_str(payload.get("supplier"))
    tonnage = _as_float(payload.get("tonnage"))
    price = _as_float(payload.get("price"))

    if not isinstance(purchase_date, dt.date):
        purchase_date = _as_date(purchase_date)
    if not isinstance(delivery_date, dt.date):
        delivery_date = _as_date(delivery_date)
    if delivery_date is not None:
        delivery_date = delivery_date.replace(day=1)

    if purchase_date is None:
        raise FeedContractError("Purchase date is required.")
    if delivery_date is None:
        raise FeedContractError("Delivery date is required.")
    if not product:
        raise FeedContractError("Product is required.")
    if not supplier:
        raise FeedContractError("Supplier is required.")
    if tonnage is None or tonnage < 0:
        raise FeedContractError("Tonnage must be a non-negative number.")
    if price is None or price < 0:
        raise FeedContractError("Price must be a non-negative number.")

    row.purchase_date = purchase_date
    row.delivery_date = delivery_date
    row.product = product
    row.product_type = product_type
    row.supplier = supplier
    row.tonnage = tonnage
    row.price = price
    db.commit()
    db.refresh(row)
    _ensure_options_include(
        db,
        product=product,
        product_type=product_type,
        supplier=supplier,
    )
    return {"row": row.to_dict()}


def _month_start(day: dt.date) -> dt.date:
    return day.replace(day=1)


def _month_end(day: dt.date) -> dt.date:
    last = calendar.monthrange(day.year, day.month)[1]
    return day.replace(day=last)


def _iter_months(start: dt.date, end: dt.date) -> list[dt.date]:
    start = _month_start(start)
    end = _month_start(end)
    if end < start:
        start, end = end, start
    months: list[dt.date] = []
    cursor = start
    while cursor <= end:
        months.append(cursor)
        if cursor.month == 12:
            cursor = dt.date(cursor.year + 1, 1, 1)
        else:
            cursor = dt.date(cursor.year, cursor.month + 1, 1)
    return months


def default_summary_date_range(db: Session) -> dict[str, str | None]:
    """Default: one month ago → latest future delivery month with purchases."""
    today = dt.date.today()
    first_this = today.replace(day=1)
    date_from = (
        dt.date(first_this.year - 1, 12, 1)
        if first_this.month == 1
        else dt.date(first_this.year, first_this.month - 1, 1)
    )

    earliest = db.scalar(select(func.min(FeedContract.delivery_date)))
    latest = db.scalar(select(func.max(FeedContract.delivery_date)))
    date_to = _month_start(today)
    if latest is not None and _month_start(latest) > date_to:
        date_to = _month_start(latest)

    bounds_min = _month_start(earliest) if earliest else date_from
    bounds_max = _month_start(latest) if latest else date_to
    if date_from < bounds_min:
        bounds_min = date_from
    if date_to > bounds_max:
        bounds_max = date_to

    return {
        "date_from": date_from.isoformat()[:7],
        "date_to": date_to.isoformat()[:7],
        "bounds_min": bounds_min.isoformat()[:7],
        "bounds_max": bounds_max.isoformat()[:7],
    }


def feed_contracts_summary(
    db: Session,
    *,
    date_from: dt.date | None = None,
    date_to: dt.date | None = None,
) -> dict[str, Any]:
    """Month × product pivot: tonnes sum and tonnage-weighted average price."""
    defaults = default_summary_date_range(db)
    if date_from is None:
        date_from = dt.date.fromisoformat(defaults["date_from"] + "-01")
    if date_to is None:
        date_to = _month_end(dt.date.fromisoformat(defaults["date_to"] + "-01"))
    else:
        date_to = _month_end(date_to)
    date_from = _month_start(date_from)

    rows = list(
        db.scalars(
            select(FeedContract)
            .where(FeedContract.delivery_date >= date_from)
            .where(FeedContract.delivery_date <= date_to)
            .order_by(
                FeedContract.delivery_date.asc(),
                FeedContract.product.asc(),
                FeedContract.id.asc(),
            )
        ).all()
    )

    # product -> type (prefer configured type from rows)
    product_type_map: dict[str, str] = {}
    aggregates: dict[tuple[str, str], dict[str, float]] = {}
    details: dict[str, list[dict[str, Any]]] = {}

    for row in rows:
        product = row.product
        ptype = row.product_type or "Other"
        if product not in product_type_map and row.product_type:
            product_type_map[product] = row.product_type
        elif product not in product_type_map:
            product_type_map[product] = ptype

        month_key = _month_start(row.delivery_date).isoformat()[:7]
        agg_key = (month_key, product)
        bucket = aggregates.setdefault(
            agg_key, {"tonnage": 0.0, "price_tonnage": 0.0}
        )
        tonnage = float(row.tonnage or 0)
        price = float(row.price or 0)
        bucket["tonnage"] += tonnage
        bucket["price_tonnage"] += price * tonnage

        details.setdefault(month_key, []).append(
            {
                "id": row.id,
                "purchase_date": row.purchase_date.isoformat(),
                "delivery_date": row.delivery_date.isoformat(),
                "product": product,
                "product_type": row.product_type,
                "tonnage": row.tonnage,
                "price": row.price,
                "supplier": row.supplier,
            }
        )

    # Column order: group by type, then product name.
    type_order = list(get_option_list(db, "product_types"))
    for ptype in sorted({product_type_map[p] for p in product_type_map}):
        if ptype not in type_order:
            type_order.append(ptype)

    products_by_type: dict[str, list[str]] = {t: [] for t in type_order}
    for product in sorted(product_type_map.keys(), key=str.casefold):
        ptype = product_type_map[product]
        products_by_type.setdefault(ptype, []).append(product)
    # Drop empty type groups.
    columns = [
        {"product_type": ptype, "products": products}
        for ptype, products in products_by_type.items()
        if products
    ]

    months = _iter_months(date_from, date_to)
    month_rows: list[dict[str, Any]] = []
    for month in months:
        month_key = month.isoformat()[:7]
        cells: dict[str, dict[str, float | None]] = {}
        for column in columns:
            for product in column["products"]:
                bucket = aggregates.get((month_key, product))
                if not bucket or bucket["tonnage"] <= 0:
                    cells[product] = {"tonnage": None, "price": None}
                else:
                    cells[product] = {
                        "tonnage": round(bucket["tonnage"], 2),
                        "price": round(
                            bucket["price_tonnage"] / bucket["tonnage"], 2
                        ),
                    }
        month_rows.append(
            {
                "month": month_key,
                "cells": cells,
                "details": details.get(month_key, []),
            }
        )

    return {
        "date_from": date_from.isoformat()[:7],
        "date_to": date_to.isoformat()[:7],
        "defaults": defaults,
        "columns": columns,
        "months": month_rows,
    }


def _distinct_contract_values(db: Session, attr: str) -> list[str]:
    column = getattr(FeedContract, attr)
    values = {
        str(v).strip()
        for v in db.scalars(select(column).distinct()).all()
        if v and str(v).strip()
    }
    return sorted(values, key=str.casefold)


def _load_option_list(db: Session, setting_key: str) -> list[str] | None:
    row = db.scalar(select(AppSetting).where(AppSetting.key == setting_key))
    if row is None or not (row.value or "").strip():
        return None
    try:
        parsed = json.loads(row.value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return None
    if not isinstance(parsed, list):
        return None
    return [str(item).strip() for item in parsed if str(item).strip()]


def _save_option_list(db: Session, setting_key: str, values: list[str]) -> None:
    payload = json.dumps(values, ensure_ascii=False)
    row = db.scalar(select(AppSetting).where(AppSetting.key == setting_key))
    if row is None:
        db.add(AppSetting(key=setting_key, value=payload))
    else:
        row.value = payload
    db.commit()


def get_option_list(db: Session, kind: str) -> list[str]:
    if kind not in _OPTION_KEYS:
        raise FeedContractError("Unknown option list.")
    setting_key = _OPTION_KEYS[kind]
    existing = _load_option_list(db, setting_key)
    if existing is not None:
        return existing

    if kind == "product_types":
        seeded = list(FEED_PRODUCT_TYPES_DEFAULT)
        from_data = _distinct_contract_values(db, "product_type")
        merged = sorted(set(seeded) | set(from_data), key=str.casefold)
    elif kind == "products":
        merged = _distinct_contract_values(db, "product")
    else:
        merged = _distinct_contract_values(db, "supplier")
    _save_option_list(db, setting_key, merged)
    return merged


def get_feed_contract_options(db: Session) -> dict[str, list[str]]:
    return {
        "products": get_option_list(db, "products"),
        "product_types": get_option_list(db, "product_types"),
        "suppliers": get_option_list(db, "suppliers"),
    }


def add_feed_option(db: Session, kind: str, value: str) -> list[str]:
    cleaned = (value or "").strip()
    if not cleaned:
        raise FeedContractError("Value cannot be empty.")
    if len(cleaned) > 128:
        raise FeedContractError("Value is too long.")
    values = get_option_list(db, kind)
    if any(existing.casefold() == cleaned.casefold() for existing in values):
        raise FeedContractError(f'"{cleaned}" already exists.')
    values.append(cleaned)
    values.sort(key=str.casefold)
    _save_option_list(db, _OPTION_KEYS[kind], values)
    return values


def remove_feed_option(db: Session, kind: str, value: str) -> list[str]:
    cleaned = (value or "").strip()
    values = get_option_list(db, kind)
    remaining = [v for v in values if v.casefold() != cleaned.casefold()]
    if len(remaining) == len(values):
        raise FeedContractError(f'"{cleaned}" was not found.')
    _save_option_list(db, _OPTION_KEYS[kind], remaining)
    return remaining


def _ensure_options_include(
    db: Session,
    *,
    product: str | None,
    product_type: str | None,
    supplier: str | None,
) -> None:
    for kind, value in (
        ("products", product),
        ("product_types", product_type),
        ("suppliers", supplier),
    ):
        if not value:
            continue
        values = get_option_list(db, kind)
        if any(v.casefold() == value.casefold() for v in values):
            continue
        values.append(value)
        values.sort(key=str.casefold)
        _save_option_list(db, _OPTION_KEYS[kind], values)


def seed_feed_contracts_if_empty(db: Session) -> dict[str, Any]:
    """Load app/seed_data/feedcontracts.xlsx when the table has no rows."""
    existing = db.scalar(select(func.count()).select_from(FeedContract)) or 0
    if existing > 0:
        # Still ensure option lists exist for the gear UI.
        get_feed_contract_options(db)
        return {"seeded": 0, "skipped": True, "reason": "table_not_empty"}

    if not _SEED_PATH.is_file():
        logger.warning("Feed contracts seed file missing: %s", _SEED_PATH)
        return {"seeded": 0, "skipped": True, "reason": "file_missing"}

    parsed = parse_feedcontracts_xlsx(_SEED_PATH.read_bytes())
    if not parsed:
        return {"seeded": 0, "skipped": True, "reason": "no_rows"}

    db.bulk_insert_mappings(
        FeedContract,
        [
            {
                **row,
                "source_file": SEED_SOURCE_FILE,
            }
            for row in parsed
        ],
    )
    db.commit()
    get_feed_contract_options(db)
    logger.info("Seeded %s feed contracts from %s", len(parsed), SEED_SOURCE_FILE)
    return {"seeded": len(parsed), "skipped": False}
