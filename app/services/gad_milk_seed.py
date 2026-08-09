"""Seed historical GAD milk tickets from app/seed_data/gadmilk.xlsx.

The workbook is a simple daily sheet: Date, Load 1, Load2, Cows In Milk.
Rows with no volumes are ignored (future placeholders). Existing GAD
collections for a date are left untouched so email imports and edits win.
"""

from __future__ import annotations

import datetime as dt
import io
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import MilkCollection
from app.services.haulier_collections import (
    SEED_GAD_MILK_SOURCE_FILE,
    _MANUAL_LOAD_ARRIVALS,
)

_SEED_PATH = Path(__file__).resolve().parent.parent / "seed_data" / "gadmilk.xlsx"
_FARM = "GAD"


def parse_gadmilk_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Parse gadmilk.xlsx into one dict per day that has at least one load."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        days: list[dict[str, Any]] = []
        for index, row in enumerate(ws.iter_rows(values_only=True)):
            if index == 0:
                continue
            raw_date = row[0] if row else None
            if raw_date is None:
                continue
            if isinstance(raw_date, dt.datetime):
                day = raw_date.date()
            elif isinstance(raw_date, dt.date):
                day = raw_date
            else:
                continue

            loads: list[int] = []
            for col in (1, 2):
                value = row[col] if col < len(row) else None
                if value is None or value == "":
                    continue
                try:
                    volume = int(round(float(value)))
                except (TypeError, ValueError):
                    continue
                if volume < 0:
                    continue
                loads.append(volume)
            if not loads:
                continue

            cows: int | None = None
            cows_raw = row[3] if len(row) > 3 else None
            if cows_raw is not None and cows_raw != "":
                try:
                    cows = int(round(float(cows_raw)))
                except (TypeError, ValueError):
                    cows = None

            days.append(
                {
                    "collection_date": day,
                    "loads": loads,
                    "cows_in_milk": cows,
                }
            )
        return days
    finally:
        wb.close()


def seed_gad_milk_collections(db: Session, *, path: Path | None = None) -> dict[str, Any]:
    """Insert missing GAD ticket days from the seed workbook.

    Skips any date that already has one or more GAD ``MilkCollection`` rows.
    """
    seed_path = path or _SEED_PATH
    if not seed_path.is_file():
        return {
            "seed_file": str(seed_path),
            "days_in_file": 0,
            "days_inserted": 0,
            "loads_inserted": 0,
            "skipped_existing": 0,
        }

    days = parse_gadmilk_xlsx(seed_path.read_bytes())
    if not days:
        return {
            "seed_file": str(seed_path),
            "days_in_file": 0,
            "days_inserted": 0,
            "loads_inserted": 0,
            "skipped_existing": 0,
        }

    existing_dates = set(
        db.scalars(
            select(MilkCollection.collection_date).where(MilkCollection.farm == _FARM)
        ).all()
    )

    now = dt.datetime.now(dt.timezone.utc).replace(tzinfo=None)
    days_inserted = 0
    loads_inserted = 0
    skipped = 0
    for day in days:
        collection_date: dt.date = day["collection_date"]
        if collection_date in existing_dates:
            skipped += 1
            continue
        cows = day.get("cows_in_milk")
        for index, volume in enumerate(day["loads"][:3]):
            arrival = _MANUAL_LOAD_ARRIVALS[index]
            db.add(
                MilkCollection(
                    farm=_FARM,
                    collection_date=collection_date,
                    sample_id=None,
                    driver=None,
                    vehicle_reg=None,
                    arrival_time=arrival,
                    depart_time=None,
                    volume_litres=volume,
                    temp_c=None,
                    temp_raw=None,
                    cows_in_milk=cows,
                    source_message_id=f"seed:gadmilk:{collection_date.isoformat()}",
                    source_file=SEED_GAD_MILK_SOURCE_FILE,
                    source_received=now,
                )
            )
            loads_inserted += 1
        existing_dates.add(collection_date)
        days_inserted += 1

    if days_inserted:
        db.commit()

    return {
        "seed_file": str(seed_path),
        "days_in_file": len(days),
        "days_inserted": days_inserted,
        "loads_inserted": loads_inserted,
        "skipped_existing": skipped,
    }
