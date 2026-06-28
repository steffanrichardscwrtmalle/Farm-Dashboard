"""Pending genomic results worklist (animals submitted, awaiting results)."""

from __future__ import annotations

import datetime as dt
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import AppSetting, GenomicResult, HerdInventory
from app.services.events_common import normalize_farms
from app.services.genomic_import import normalize_hbn
from app.services.herd_import_utils import BEEF_CBREED_MIN

PENDING_RESULTS_RECIPIENT_KEY = "pending_results_recipient"
FEMALE_GENDER = "Female"

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

EMAIL_BODY = (
    "Please see attached the updated file containing the remaining genomic "
    "submissions that we are waiting for results on.\n\n"
    "Regards,\n"
    "Steffan Richards"
)

_XLSX_HEADERS = (
    "ID",
    "Ear Tag",
    "TSU No",
    "Submission Date",
    "Days Since Submission",
    "Farm",
)


def _format_date(value: str | None) -> str:
    if not value:
        return ""
    try:
        return dt.date.fromisoformat(value).strftime("%d/%m/%Y")
    except ValueError:
        return value


def list_pending_results(
    db: Session,
    *,
    farms: list[str] | None = None,
    min_days: int | None = None,
) -> dict[str, Any]:
    """Female dairy heifers (CBRD < 102) that have been TSU-tested and submitted.

    When ``min_days`` is given, only animals whose days-since-submission is at least
    that value are returned (e.g. min_days=30 hides anything submitted < 30 days ago).
    """
    selected_farms = normalize_farms(farms)
    if not selected_farms:
        return {"rows": [], "total": 0}

    today = dt.date.today()
    # Animals whose genomic results have already come back (matched on HBN = digits-only ETAG).
    genomic_hbns = set(db.scalars(select(GenomicResult.hbn)).all())

    query = (
        select(
            HerdInventory.farm,
            HerdInventory.cow_id,
            HerdInventory.etag,
            HerdInventory.gid,
            HerdInventory.subd,
        )
        .where(HerdInventory.farm.in_(selected_farms))
        .where(HerdInventory.gender == FEMALE_GENDER)
        .where(HerdInventory.cbrd.isnot(None))
        .where(HerdInventory.cbrd < BEEF_CBREED_MIN)
        .where(HerdInventory.subd.isnot(None))
        .where(HerdInventory.gtest.isnot(None))
        .order_by(HerdInventory.subd.asc(), HerdInventory.etag.asc())
    )

    threshold = min_days if (min_days is not None and min_days > 0) else None

    rows: list[dict[str, Any]] = []
    for farm, cow_id, etag, gid, subd in db.execute(query).all():
        # Exclude animals that already have genomic results.
        hbn = normalize_hbn(etag)
        if hbn and hbn in genomic_hbns:
            continue
        days = (today - subd).days if subd else None
        if threshold is not None and (days is None or days < threshold):
            continue
        rows.append(
            {
                "id": (cow_id or "").strip(),
                "etag": (etag or "").strip(),
                "gid": (gid or "").strip(),
                "subd": subd.isoformat() if subd else "",
                "days_since_submission": days,
                "farm": farm,
            }
        )

    return {"rows": rows, "total": len(rows)}


def build_pending_results_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Results"
    ws.append(list(_XLSX_HEADERS))
    for row in rows:
        ws.append(
            [
                row.get("id", ""),
                row.get("etag", ""),
                row.get("gid", ""),
                _format_date(row.get("subd")),
                row.get("days_since_submission"),
                row.get("farm", ""),
            ]
        )

    widths = [12, 18, 14, 16, 22, 8]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def get_recipient(db: Session) -> dict[str, str | None]:
    row = db.scalar(
        select(AppSetting).where(AppSetting.key == PENDING_RESULTS_RECIPIENT_KEY)
    )
    return {"recipient": row.value if row else None}


def set_recipient(db: Session, email: str) -> dict[str, str]:
    normalized = (email or "").strip()
    if not normalized:
        raise ValueError("Recipient email is required.")
    row = db.scalar(
        select(AppSetting).where(AppSetting.key == PENDING_RESULTS_RECIPIENT_KEY)
    )
    if row is None:
        row = AppSetting(key=PENDING_RESULTS_RECIPIENT_KEY, value=normalized)
        db.add(row)
    else:
        row.value = normalized
    db.commit()
    return {"recipient": normalized}
