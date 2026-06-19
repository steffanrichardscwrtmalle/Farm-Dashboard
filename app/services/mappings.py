"""Category and product mapping rules from keywords.xlsx or the database."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import ProductMappingRule
from app.services.mapping_options import ensure_mapping_option

_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_KEYWORDS_PATH = _PROJECT_ROOT / "keywords.xlsx"
SHEET_NAME = "Combined"


def load_rules_from_excel_bytes(file_bytes: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    try:
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
        return _parse_worksheet(ws)
    finally:
        wb.close()


def load_rules_from_excel_path(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
        return _parse_worksheet(ws)
    finally:
        wb.close()


def _parse_worksheet(ws) -> list[dict[str, str]]:
    headers = _header_map(ws)
    keyword_col = headers.get("keyword", 1)
    category_col = headers.get("category", 2)
    farm_col = headers.get("farm description", 3)

    rules: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        keyword = _cell_str(ws.cell(row=row_idx, column=keyword_col).value)
        if not keyword:
            continue
        category = _cell_str(ws.cell(row=row_idx, column=category_col).value)
        farm = _cell_str(ws.cell(row=row_idx, column=farm_col).value)
        rules.append(
            {
                "keyword": keyword,
                "category": category,
                "farm_description": farm,
            }
        )
    return rules


def _header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is None:
            if headers:
                break
            continue
        headers[str(val).strip().lower()] = col_idx
    return headers


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_product_mapping_rules(db: Session) -> list[tuple[str, str, str]]:
    """Return (keyword_lower, category, farm_description) ordered for first-match wins."""
    rules = db.scalars(
        select(ProductMappingRule).order_by(ProductMappingRule.sort_order, ProductMappingRule.id)
    ).all()
    return [
        (r.keyword.lower(), r.category, r.farm_description)
        for r in rules
        if r.keyword.strip()
    ]


def import_rules_to_db(db: Session, rule_dicts: list[dict[str, str]], *, replace: bool = True) -> int:
    if replace:
        for rule in db.scalars(select(ProductMappingRule)).all():
            db.delete(rule)
        db.flush()

    for i, row in enumerate(rule_dicts):
        category = row.get("category") or ""
        farm = row.get("farm_description") or ""
        ensure_mapping_option(db, "category", category)
        ensure_mapping_option(db, "farm_description", farm)
        db.add(
            ProductMappingRule(
                keyword=row["keyword"],
                category=category,
                farm_description=farm,
                sort_order=i,
            )
        )
    db.commit()
    return len(rule_dicts)


def seed_mappings_if_empty(db: Session, keywords_path: Path | None = None) -> int:
    if db.scalars(select(ProductMappingRule).limit(1)).first() is not None:
        return 0
    path = keywords_path or DEFAULT_KEYWORDS_PATH
    rules = load_rules_from_excel_path(path)
    if not rules:
        return 0
    return import_rules_to_db(db, rules, replace=True)


def list_mapping_rules(db: Session) -> list[ProductMappingRule]:
    return list(
        db.scalars(
            select(ProductMappingRule).order_by(ProductMappingRule.sort_order, ProductMappingRule.id)
        ).all()
    )
