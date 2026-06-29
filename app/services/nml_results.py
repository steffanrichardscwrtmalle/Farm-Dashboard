"""Query and export NML milk-quality results for the dashboard page."""

from __future__ import annotations

import csv
import datetime as dt
import io
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import NmlMilkResult
from app.services.events_common import normalize_farms

XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

_HEADERS = (
    "Farm",
    "Producer Ref",
    "Milk Buyer",
    "Sample Date",
    "Sample ID",
    "Butterfat %",
    "Protein %",
    "SCC",
    "BactoScan",
    "FPD",
    "A/B",
    "Urea %",
)

# Metrics averaged per day for the trend charts.
_TREND_METRICS = ("butterfat_pct", "protein_pct", "scc", "bactoscan", "urea_pct")


def _ab_label(value: bool | None) -> str:
    if value is True:
        return "Pass"
    if value is False:
        return "Fail"
    return ""


def _row_to_dict(row: NmlMilkResult) -> dict[str, Any]:
    return {
        "farm": row.farm or "",
        "producer_ref": row.producer_ref or "",
        "milk_buyer": row.milk_buyer or "",
        "sample_date": row.sample_date.isoformat() if row.sample_date else "",
        "sample_id": row.sample_id or "",
        "butterfat_pct": row.butterfat_pct,
        "protein_pct": row.protein_pct,
        "scc": row.scc,
        "bactoscan": row.bactoscan,
        "fpd": row.fpd,
        "antibiotic_pass": row.antibiotic_pass,
        "urea_pct": row.urea_pct,
        "report_month": row.report_month or "",
    }


def _build_trend(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    """Daily averages per farm for charting, oldest first."""
    buckets: dict[tuple[str, str], dict[str, list[float]]] = {}
    for row in rows:
        farm = row["farm"] or "?"
        date = row["sample_date"]
        if not date:
            continue
        bucket = buckets.setdefault((farm, date), {m: [] for m in _TREND_METRICS})
        for metric in _TREND_METRICS:
            value = row.get(metric)
            if value is not None:
                bucket[metric].append(float(value))

    trend: dict[str, list[dict[str, Any]]] = {}
    for (farm, date), metrics in buckets.items():
        point: dict[str, Any] = {"date": date}
        for metric in _TREND_METRICS:
            values = metrics[metric]
            point[metric] = round(sum(values) / len(values), 3) if values else None
        trend.setdefault(farm, []).append(point)
    for points in trend.values():
        points.sort(key=lambda item: item["date"])
    return trend


def _summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    def avg(metric: str) -> float | None:
        values = [float(r[metric]) for r in rows if r.get(metric) is not None]
        return round(sum(values) / len(values), 2) if values else None

    latest = max((r["sample_date"] for r in rows if r["sample_date"]), default="")
    fails = sum(1 for r in rows if r.get("antibiotic_pass") is False)
    return {
        "count": len(rows),
        "latest_sample_date": latest,
        "avg_butterfat_pct": avg("butterfat_pct"),
        "avg_protein_pct": avg("protein_pct"),
        "avg_scc": avg("scc"),
        "avg_bactoscan": avg("bactoscan"),
        "antibiotic_fails": fails,
    }


def list_nml_results(
    db: Session,
    *,
    farms: list[str] | None = None,
    date_from: dt.date | None = None,
    date_to: dt.date | None = None,
) -> dict[str, Any]:
    selected_farms = normalize_farms(farms)
    if not selected_farms:
        return {"rows": [], "total": 0, "summary": _summary([]), "trend": {}}

    query = select(NmlMilkResult).where(NmlMilkResult.farm.in_(selected_farms))
    if date_from is not None:
        query = query.where(NmlMilkResult.sample_date >= date_from)
    if date_to is not None:
        query = query.where(NmlMilkResult.sample_date <= date_to)
    query = query.order_by(
        NmlMilkResult.sample_date.desc(), NmlMilkResult.sample_id.desc()
    )

    rows = [_row_to_dict(row) for row in db.scalars(query).all()]
    return {
        "rows": rows,
        "total": len(rows),
        "summary": _summary(rows),
        "trend": _build_trend(rows),
    }


def _export_cells(row: dict[str, Any]) -> list[Any]:
    return [
        row.get("farm", ""),
        row.get("producer_ref", ""),
        row.get("milk_buyer", ""),
        row.get("sample_date", ""),
        row.get("sample_id", ""),
        row.get("butterfat_pct"),
        row.get("protein_pct"),
        row.get("scc"),
        row.get("bactoscan"),
        row.get("fpd"),
        _ab_label(row.get("antibiotic_pass")),
        row.get("urea_pct"),
    ]


def build_nml_results_csv(rows: list[dict[str, Any]]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(_HEADERS)
    for row in rows:
        writer.writerow(_export_cells(row))
    return buffer.getvalue()


def build_nml_results_xlsx(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "NML Results"
    ws.append(list(_HEADERS))
    for row in rows:
        ws.append(_export_cells(row))

    widths = [8, 14, 16, 14, 12, 12, 11, 8, 11, 8, 8, 9]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
