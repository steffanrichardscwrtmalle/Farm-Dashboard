"""
Read-only migration from desktop Wynnstay-Invoices Excel files into the web app database.
Never writes back to the desktop Tool Files.
"""

from __future__ import annotations

import argparse
import datetime
import sys
from pathlib import Path

from openpyxl import load_workbook

# Allow running as: python scripts/migrate_from_excel.py
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_PROJECT_ROOT))

from sqlalchemy import func, select

from app.db import SessionLocal, init_db  # noqa: E402
from app.models import DEFAULT_BUSINESS, InvoiceLine  # noqa: E402
from app.services.invoice_ops import refresh_all_invoice_lines  # noqa: E402
from app.services.mappings import seed_mappings_if_empty  # noqa: E402
from app.services.transforms import cell_to_date, is_credit_goods_value  # noqa: E402

INVOICE_SHEET_NAME = "Wynnstay Invoice Data"
TABLE_HEADER_ROW = 1
TABLE_FIRST_DATA_ROW = 2


def _header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=TABLE_HEADER_ROW, column=col_idx).value
        if val is None:
            if headers:
                break
            continue
        headers[str(val).strip().lower()] = col_idx
    return headers


def migrate_main_workbook(db, main_path: Path) -> int:
    if not main_path.exists():
        print(f"Main workbook not found: {main_path}")
        return 0

    wb = load_workbook(main_path, data_only=True)
    try:
        if INVOICE_SHEET_NAME not in wb.sheetnames:
            print(f"Sheet '{INVOICE_SHEET_NAME}' not found.")
            return 0
        ws = wb[INVOICE_SHEET_NAME]
        headers = _header_map(ws)
        if not headers:
            print("No headers found.")
            return 0

        last_row = ws.max_row
        count = 0
        for row_idx in range(TABLE_FIRST_DATA_ROW, last_row + 1):
            def cell(name: str):
                col = headers.get(name.lower())
                if col is None:
                    return None
                return ws.cell(row=row_idx, column=col).value

            if all(cell(k) in (None, "") for k in headers):
                continue

            goods_value = _float(cell("goods value"))
            line = InvoiceLine(
                business=DEFAULT_BUSINESS,
                date=cell_to_date(cell("date")),
                reference=_str(cell("reference")),
                product_code=_str(cell("product code")),
                category=_str(cell("category")),
                product_description=_str(cell("product description")),
                farm_description=_str(cell("farm description")),
                quantity=_float(cell("quantity")),
                unit=_str(cell("unit")),
                price=_float(cell("price")),
                goods_value=goods_value,
                vat=_float(cell("vat")),
                total=_float(cell("total")),
                date_added=cell_to_date(cell("date added")),
                invoice_date=cell_to_date(cell("invoice date")),
                recent=_str(cell("recent")),
                credit="Yes" if is_credit_goods_value(goods_value) else "No",
            )
            db.add(line)
            count += 1

        db.commit()
        return count
    finally:
        wb.close()


def _str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def _float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def main() -> None:
    parser = argparse.ArgumentParser(description="Migrate desktop Excel data into web app DB (read-only).")
    parser.add_argument(
        "--desktop-dir",
        type=Path,
        default=_PROJECT_ROOT.parent / "Wynnstay-Invoices",
        help="Path to the desktop Wynnstay-Invoices folder",
    )
    args = parser.parse_args()

    tool_files = args.desktop_dir / "Tool Files"
    main_workbook = tool_files / "WynnstayInvoices.xlsx"
    category_map = tool_files / "Wynnstay Category Map.xlsx"

    init_db()
    db = SessionLocal()
    try:
        existing = db.scalar(select(func.count()).select_from(InvoiceLine)) or 0
        if existing:
            print(f"Database already has {existing} invoice lines. Skipping row import.")
        else:
            n = migrate_main_workbook(db, main_workbook)
            print(f"Imported {n} invoice lines from {main_workbook.name}")

        n_rules = seed_mappings_if_empty(db)
        print(f"Seeded {n_rules} keyword mapping rules from keywords.xlsx")

        refreshed = refresh_all_invoice_lines(db)
        print(f"Refreshed {refreshed} rows after migration.")
    finally:
        db.close()


if __name__ == "__main__":
    main()
