"""One-off: import HP schedules from Book2.xlsx into local SQLite."""

from __future__ import annotations

import datetime as dt
import shutil
import sys
import tempfile
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from sqlalchemy import select

from app.db import SessionLocal, init_db
from app.models import HpSchedule
from app.services.hp_schedules import create_hp_schedule, list_hp_schedules

SRC = ROOT / "Book2.xlsx"


def _as_date(value) -> dt.date:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    raise ValueError(f"Expected date, got {value!r}")


def load_agreements(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    groups: dict[str, list[dict]] = defaultdict(list)

    for r in range(2, (ws.max_row or 1) + 1):
        raw_name = ws.cell(r, 1).value
        if not raw_name:
            continue
        name = str(raw_name).replace("\r\n", "\n").strip()
        pay_date = ws.cell(r, 6).value
        if not pay_date:
            continue
        groups[name].append(
            {
                "description": str(ws.cell(r, 2).value or "").strip(),
                "capital": float(ws.cell(r, 3).value or 0),
                "interest": float(ws.cell(r, 4).value or 0),
                "date": _as_date(pay_date),
                "day": int(ws.cell(r, 7).value or _as_date(pay_date).day),
            }
        )

    agreements: list[dict] = []
    for name, rows in groups.items():
        rows_sorted = sorted(rows, key=lambda x: x["date"])
        first = rows_sorted[0]
        # Prefer the most common monthly amounts if a few rows differ.
        caps = [round(x["capital"], 2) for x in rows_sorted]
        ints = [round(x["interest"], 2) for x in rows_sorted]
        monthly_capital = max(set(caps), key=caps.count)
        monthly_interest = max(set(ints), key=ints.count)
        days = [x["day"] for x in rows_sorted]
        payment_day = max(set(days), key=days.count)
        descriptions = [x["description"] for x in rows_sorted if x["description"]]
        description = descriptions[0] if descriptions else ""

        agreements.append(
            {
                "name": name.replace("\n", " ").strip(),
                "description": description,
                "monthly_capital": monthly_capital,
                "monthly_interest": monthly_interest,
                "months": len(rows_sorted),
                "payment_day": payment_day,
                "start_month": first["date"].replace(day=1),
                "first_payment": first["date"],
                "last_payment": rows_sorted[-1]["date"],
            }
        )

    agreements.sort(key=lambda a: (a["payment_day"], a["name"].lower()))
    return agreements


def main() -> None:
    if not SRC.exists():
        raise SystemExit(f"Missing {SRC}")

    # Copy in case Excel has the file locked.
    with tempfile.TemporaryDirectory() as tmp:
        copy = Path(tmp) / "Book2.xlsx"
        shutil.copy2(SRC, copy)
        agreements = load_agreements(copy)

    print(f"Parsed {len(agreements)} HP agreements from {SRC.name}")
    for a in agreements:
        print(
            f"  {a['months']:3}m  {a['first_payment']} -> {a['last_payment']}  "
            f"day={a['payment_day']:2}  "
            f"cap={a['monthly_capital']:,.2f}  int={a['monthly_interest']:,.2f}  "
            f"{a['description'][:28]:28}  {a['name'][:55]}"
        )

    init_db()
    db = SessionLocal()
    try:
        # Replace existing active schedules so re-import is clean.
        existing = db.scalars(select(HpSchedule)).all()
        for row in existing:
            row.is_active = False
        db.commit()

        created = 0
        for a in agreements:
            create_hp_schedule(
                db,
                name=a["name"],
                business="CM",
                description=a["description"],
                monthly_capital=a["monthly_capital"],
                monthly_interest=a["monthly_interest"],
                months=a["months"],
                payment_day=a["payment_day"],
                start_month=a["start_month"],
                user_id=None,
            )
            created += 1

        listed = list_hp_schedules(db)
        print(f"\nImported {created} schedules. Active in DB: {len(listed)}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
